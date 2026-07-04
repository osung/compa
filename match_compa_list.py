# -*- coding: utf-8 -*-
"""compa_list.xlsx 의 각 기업 수요에 대해 임베딩 매칭 + LLM 재랭킹으로 과제 추천.

파이프라인 (기업 수요 1건 = xlsx 한 행)
  1) 임베딩 매칭으로 후보 Top-K(기본 100) 선별
       - 기업명으로 회사 키워드 리스트 조회(company_match_data_260612_emb.pkl)
       - 수요기술명/수요기술 내용/예상 적용 제품 및 서비스 를 normalize_user_input 으로
         명사 키워드 추출해 합침 → pro-sroberta 인코딩 → 과제 norm_embed 와 코사인
       - 매칭점수 = 0.6*cos + 0.4*clip(유망성/100) (matching_viewer 기본) 로 Top-K
  2) LLM 재랭킹: 후보 Top-K 의 (과제명·과제설명문·키워드)를 읽고, 수요기술명/내용/
     사양/예상 적용 제품 및 서비스 에 대한 적합도를 0~100 으로 점수화(배치 처리).
       - 백엔드: matching_viewer_explain 자동 선택(mac=MLX / CUDA=vLLM)
  3) LLM 점수 내림차순 Top-N(기본 10) 추천 → 출력 xlsx 새 탭에 저장.

출력 (OUT_XLSX, 원본 compa_list.xlsx 는 수정하지 않음)
  - 'LLM_추천_top10' : LLM 점수 상위 10 (최종 추천)
  - '매칭결과'        : 임베딩 매칭 상위 10 (참고용)

진행 체크포인트(LLM_CKPT)에 (번호, 과제고유번호)별 점수를 저장해 재실행 시 이어서 진행.
"""
import os
import re
import json
import pickle
import argparse

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer

import user_input_keywords as uik
import matching_viewer_explain as mve

# ---- 설정 ----------------------------------------------------------------
XLSX = "compa_list.xlsx"              # 입력(원본, 수정하지 않음)
OUT_XLSX = "compa_list_매칭결과.xlsx"  # 출력
EMB_SHEET = "매칭결과"                 # 임베딩 Top10 (참고)
LLM_SHEET = "LLM_추천_top10"           # LLM 재랭킹 Top10 (최종)
LLM_CKPT = "llm_scores_checkpoint.json"

COMPANY_DATA = "company_match_data_260612_emb.pkl"
PROJECT_EMB = "public_RnD_embeddings_pro_260601_with_desc.pkl"
PROJECT_META = "project_match_data_260612.pkl"   # 과제수행기관명 조인용
MODEL_DIR = "pro-sroberta"

TOPK_CAND = 100        # 임베딩 후보 수 (LLM 평가 대상)
TOPN_FINAL = 10        # 최종 추천 수
LLM_BATCH = 20         # LLM 한 번에 평가할 과제 수
LLM_MAXTOK = 3500      # 배치당 생성 토큰 상한
DESC_CAND = 280        # LLM 프롬프트용 과제설명문 절단 길이
DESC_OUT = 30000       # 엑셀 셀 길이 한계(32767) 회피용 출력 설명문 절단
SEP = ";"
W_COS, W_PROM = 0.6, 0.4
RANK_BY = "match"      # 후보 선별 기준: "match"(0.6cos+0.4유망성) | "cos"(순수 코사인)
# -------------------------------------------------------------------------


def norm_name(s):
    s = str(s)
    s = re.sub(r"\(주\)|\(株\)|㈜|주식회사|\(유\)|유한회사|\(재\)|재단법인|"
               r"\(사\)|사단법인|농업회사법인|\(농\)", "", s)
    s = re.sub(r"\s+", "", s)
    return s.strip().lower()


def dedup(tokens):
    seen, out = set(), []
    for v in tokens:
        v = str(v).strip()
        k = v.casefold()
        if v and k not in seen:
            seen.add(k)
            out.append(v)
    return out


def build_name_index(cdata):
    idx = {}
    for rec in cdata.values():
        nm = rec.get("업체명", "")
        if not nm:
            continue
        key = norm_name(nm)
        kws = rec.get("키워드_리스트") or []
        if key not in idx or len(kws) > len(idx[key]):
            idx[key] = list(kws)
    return idx


def cell(rec, col):
    v = rec.get(col)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


# ---- LLM 적합도 평가 ------------------------------------------------------
_SYS = ("당신은 기업의 기술 수요와 국가 R&D 과제의 적합성을 평가하는 기술이전 전문가입니다. "
        "각 과제가 기업의 수요기술을 해결·지원하는 데 실제로 얼마나 적합한지 냉정하게 평가합니다.")

_GUIDE = (
    "각 과제가 위 [기업 기술 수요]를 해결·지원하는 데 얼마나 적합한지 0~100 정수로 평가하세요.\n"
    "  90-100: 동일 기술/제품을 직접 해결하는 매우 높은 적합\n"
    "  70-89 : 핵심 기술·응용 분야가 상당 부분 부합\n"
    "  40-69 : 일부 요소만 관련(부분 적합)\n"
    "  10-39 : 분야만 유사하거나 약하게 관련\n"
    "  0-9   : 사실상 무관\n"
    "키워드 표면 일치가 아니라 기술 내용·적용 제품 관점의 실질 적합도를 보세요.\n"
    "반드시 아래 JSON 형식으로만, 모든 과제에 대해 답하세요(reason 은 40자 이내 한국어):\n"
    '{"results": [{"id": <과제번호>, "score": <0-100 정수>, "reason": "<근거>"}, ...]}'
)


def _demand_block(d):
    return ("[기업 기술 수요]\n"
            f"- 수요기술명: {d.get('수요기술명','')}\n"
            f"- 수요기술 내용: {d.get('수요기술 내용','')}\n"
            f"- 수요기술 사양: {d.get('수요기술 사양','')}\n"
            f"- 예상 적용 제품 및 서비스: {d.get('예상 적용 제품 및 서비스','')}")


def _cand_block(cands):
    lines = ["[평가 대상 R&D 과제 목록]"]
    for c in cands:
        kw = ", ".join(c["키워드"][:8])
        lines.append(f"{c['id']}. 과제명: {c['과제명']}\n"
                     f"   키워드: {kw}\n"
                     f"   설명: {c['과제설명문'][:DESC_CAND]}")
    return "\n".join(lines)


def _parse_scores(text):
    """LLM 출력에서 {\"results\":[...]} JSON 을 견고하게 파싱 → {id: (score, reason)}."""
    out = {}
    m = re.search(r'\{.*"results".*\}', text, re.DOTALL)
    blob = m.group(0) if m else text
    try:
        data = json.loads(blob)
        items = data.get("results", []) if isinstance(data, dict) else data
    except Exception:
        # 폴백: 개별 객체 단위 정규식 추출
        items = []
        for mm in re.finditer(
                r'\{\s*"id"\s*:\s*(\d+)\s*,\s*"score"\s*:\s*(\d+)\s*,\s*"reason"\s*:\s*"([^"]*)"',
                text):
            items.append({"id": int(mm.group(1)), "score": int(mm.group(2)),
                          "reason": mm.group(3)})
    for it in items:
        try:
            i = int(it["id"])
            s = max(0, min(100, int(round(float(it["score"])))))
            out[i] = (s, str(it.get("reason", "")).strip()[:120])
        except (KeyError, ValueError, TypeError):
            continue
    return out


def llm_score(demand, cands):
    """후보 리스트를 LLM_BATCH 단위로 평가. cands: [{id,과제명,과제설명문,키워드}], id 는 1..K.
    반환: {id: (score, reason)} (누락분은 score=0)."""
    scores = {}
    for b in range(0, len(cands), LLM_BATCH):
        batch = cands[b:b + LLM_BATCH]
        user = f"{_demand_block(demand)}\n\n{_cand_block(batch)}\n\n{_GUIDE}"
        msgs = [{"role": "system", "content": _SYS},
                {"role": "user", "content": user}]
        try:
            out = mve.stream_explanation(msgs, max_tokens=LLM_MAXTOK,
                                         temperature=0.0, top_p=1.0)
            parsed = _parse_scores(out)
        except Exception as e:
            print(f"      ! LLM 배치 실패({b}): {e}")
            parsed = {}
        for c in batch:
            scores[c["id"]] = parsed.get(c["id"], (0, "(LLM 응답 누락)"))
    return scores
# -------------------------------------------------------------------------


# ---- 상세 추천 근거 설명 (matching_viewer_explain 재사용 + 수요기술 사양 근거 추가) ----
# mve.build_messages(payload, direction='company') 는 _DIRECTIONS['company'] 의
# system 프롬프트·few-shot 을 그대로 쓰고 payload(JSON)를 근거로 추천 근거를 생성한다.
# 여기서는 수요기업↔과제 맥락에 맞춘 payload 를 직접 구성하고, 기존 3개 섹션에
# '수요기술 사양 적합성' 섹션을 추가해 수요기술 사양(정량 스펙) 충족 여부를 함께 서술하게 한다.
EXPLAIN_CKPT = "llm_explain_checkpoint.json"
EX_FMT = ["연관성", "수요기술 사양 적합성", "추천 과제의 우수성", "유사 사례 및 실적"]
_EX_GUIDE = {
    "연관성": ("company.description(해당 기업의 사업영역·보유역량·특성)과 company.수요기술명·"
            "수요기술_내용(이 기업이 '필요로 하는'(아직 보유하지 않은) 기술의 특성)을 함께 고려하여, "
            "이 기업의 특성에 비추어 왜 이 수요기술이 필요한지 맥락을 짚고, 그 수요기술과 과제의 "
            "목표·내용 사이의 기술적 연관성을 설명. "
            "수요기술은 기업이 보유한 기술이 아니라 확보하려는 기술이므로 '기업이 이미 ~기술을 "
            "보유/개발했다'거나 '~공정·~소재를 보유하고 있다'고 단정하지 말 것(수요기술 내용을 "
            "기업의 현재 사업·역량으로 서술 금지). company.description 이 비어 있거나 desc_ok=0 이면 "
            "수요기술 정보를 우선 사용하되, 기업이 보유하지 않은 역량을 임의로 가정하지 말 것"),
    "수요기술 사양 적합성": ("company.수요기술_사양의 정량·정성 요구(수치·기준·성능)를 과제가 충족 또는 "
                       "근접하는지, 어떤 항목이 부합하고 어떤 항목이 미흡/불확실한지 구체적으로 평가"),
    "추천 과제의 우수성": "과제의 유망성·연구성과(논문/특허)·수행기관 역량 등 추천 과제의 강점",
    "유사 사례 및 실적": "과제의 논문·특허 등 실적이 수요 해결에 주는 시사점(없으면 생략)",
}


def build_demand_payload(demand, proj):
    """수요기업↔과제 맥락의 mve 호환 payload 구성(수요기술 사양 포함).

    demand 에 '기업설명문'(회사 임베딩 DB 의 기업 설명문)이 있으면 이를 회사 실제
    설명문으로 보아 company.description 으로 사용한다(desc_ok/desc_issue 동반).
    수요기술 내용은 항상 별도 필드('수요기술_내용')로 전달한다 → 추천 근거가
    수요기술명/내용/사양뿐 아니라 기업 사업영역·강점까지 반영한다.

    중요: 수요기술은 '기업이 필요로 하는(아직 보유하지 않은)' 기술이다. 과거에는
    기업설명문이 없을 때 수요기술 내용을 company.description 으로 넣어, 모델이 그
    수요기술을 '기업이 이미 보유한 사업/역량'으로 오인해 서술하는 문제가 있었다.
    이를 막기 위해 기업설명문이 없으면 description 을 비워 두고(수요기술 정보는
    '수요기술_*' 필드로만 제공), 가이드(_EX_GUIDE)에서 '기업이 이미 보유했다고
    단정하지 말 것'을 명시한다."""
    comp_desc = str(demand.get("기업설명문") or "").strip()
    has_cdesc = bool(comp_desc)
    company = {
        "company_id": "", "name": demand.get("기업명", ""),
        # 기업 실제 설명문만 description 으로 사용(없으면 빈 문자열 → 수요기술을 보유역량으로 오인 방지)
        "description": comp_desc,
        "desc_issue": str(demand.get("desc_issue", "") or "") if has_cdesc else "",
        "desc_ok": int(demand.get("desc_ok", 1)) if has_cdesc else 1,
        "수요기술명": demand.get("수요기술명", ""),
        # 수요기술 내용 = 기업이 '확보하려는' 기술의 특성. 항상 별도 필드로 전달.
        "수요기술_내용": demand.get("수요기술 내용", ""),
        "수요기술_사양": demand.get("수요기술 사양", ""),
        "예상_적용_제품_및_서비스": demand.get("예상 적용 제품 및 서비스", ""),
        "keyword": list(demand.get("keywords", []))[:30],
        "purpose": [x for x in (demand.get("수요기술명", ""),
                                demand.get("예상 적용 제품 및 서비스", "")) if x],
        "patent_matching_related": [], "has_patent_matching_related": False,
        "patent_count": None, "company_patent": [],
        "conduct_list_project": [], "has_conduct_project": False, "region": "",
    }
    paper = (proj.get("논문명") or [])[:3]
    patent = (proj.get("특허명") or [])[:3]
    project = {
        "project_id": proj.get("pid", ""), "title": proj.get("과제명", ""),
        "description": (proj.get("설명", "") or "")[:1200],
        "project_score": proj.get("유망성", ""), "cosine_distance": "",
        "keyword": (proj.get("키워드") or [])[:15],
        "과제수행기관": proj.get("수행기관", ""),
        "총연구비_상위비율": proj.get("총연구비_상위비율"),
        "논문건수_상위비율": proj.get("논문건수_상위비율"),
        "특허건수_상위비율": proj.get("특허건수_상위비율"),
    }
    return {
        "company": company, "project": project,
        "conduct_list_company": [], "has_conduct_company": False, "region_relation": "",
        "related_research": {"paper": paper, "patent": patent},
        "has_paper": len(paper) > 0, "has_patent": len(patent) > 0,
        "paper_list_count": int(proj.get("논문건수") or 0),
        "patent_list_count": int(proj.get("특허건수") or 0),
        "output_requirements": {
            "language": "ko", "format": EX_FMT,
            "section_sentence_range": "각 섹션 3~4문장",
            "section_guide": _EX_GUIDE,
            "must_cover_수요기술_사양": True,
            "forbidden": ["예시", "참고", "제출", "메타"],
        },
    }


def gen_explanation(demand, proj):
    """mve 파이프라인으로 4섹션 상세 추천 근거 생성 → 한 셀 텍스트로 반환."""
    payload = build_demand_payload(demand, proj)
    msgs = mve.build_messages(payload, direction="company")
    out = mve.stream_explanation(msgs, max_tokens=1400, temperature=0.2, top_p=0.9,
                                 expected_keys=EX_FMT)
    secs = mve.parse_sections(out, tuple(EX_FMT))
    parts = [f"[{k}] {secs.get(k, '').strip()}" for k in EX_FMT if secs.get(k, "").strip()]
    return "\n\n".join(parts) if parts else out.strip()
# -------------------------------------------------------------------------


def main():
    global LLM_BATCH
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="처리할 기업 수요 수 제한(0=전체, 테스트용)")
    ap.add_argument("--topk", type=int, default=TOPK_CAND, help="임베딩 후보 수")
    ap.add_argument("--batch", type=int, default=LLM_BATCH, help="LLM 배치 크기")
    ap.add_argument("--explain", action="store_true",
                    help="최종 Top10 각 과제에 대해 mve 기반 상세 추천 근거(수요기술 사양 포함) 생성")
    args = ap.parse_args()
    LLM_BATCH = args.batch
    topk = args.topk

    print("[1/6] 회사 키워드 인덱스 로드…")
    with open(COMPANY_DATA, "rb") as f:
        cdata = pickle.load(f)
    name_idx = build_name_index(cdata)
    del cdata

    print("[2/6] xlsx 수요 리스트 로드…")
    xl = pd.read_excel(XLSX, header=2)
    xl = xl[xl["기업명"].notna()].reset_index(drop=True)
    records = xl.to_dict("records")
    if args.limit:
        records = records[:args.limit]
    print(f"      기업 수요 {len(records)}건 (전체 {len(xl)})")

    print("[3/6] 과제 코퍼스 로드…")
    df = pd.read_pickle(PROJECT_EMB)
    pid = df["과제고유번호"].astype(str).values
    pname = df["과제명"].astype(str).values
    pdesc = df["과제설명문"].astype(str).values
    pkw = df["키워드_리스트"].values
    promise = df["유망성점수"].values.astype(np.float32)
    M = np.vstack([np.asarray(e, dtype=np.float32) for e in df["norm_embed"].values])
    M /= np.linalg.norm(M, axis=1, keepdims=True)
    del df
    with open(PROJECT_META, "rb") as f:
        pmeta = pickle.load(f)
    org = np.array([str(pmeta.get(p, {}).get("과제수행기관명", "")) for p in pid])
    if not args.explain:
        del pmeta   # 상세설명(--explain)에선 논문/특허 성과 조회에 pmeta 가 필요해 보존
    pnorm = np.clip(promise / 100.0, 0, 1)

    print("[4/6] pro-sroberta 로드…")
    model = SentenceTransformer(MODEL_DIR, device="cpu")
    model.eval()

    def encode(text):
        e = model.encode([text], convert_to_numpy=True, normalize_embeddings=False,
                         show_progress_bar=False)[0].astype(np.float32)
        n = np.linalg.norm(e)
        return e / n if n else e

    # LLM 점수 체크포인트
    ckpt = {}
    if os.path.exists(LLM_CKPT):
        with open(LLM_CKPT, encoding="utf-8") as f:
            ckpt = json.load(f)
        print(f"      체크포인트 로드: {len(ckpt)}개 수요 평가 완료분 존재")
    ex_ckpt = {}
    if args.explain and os.path.exists(EXPLAIN_CKPT):
        with open(EXPLAIN_CKPT, encoding="utf-8") as f:
            ex_ckpt = json.load(f)
        print(f"      설명 체크포인트 로드: {len(ex_ckpt)}건")

    print("[5/6] 임베딩 매칭 + LLM 재랭킹…")
    emb_rows, llm_rows = [], []
    for n, rec in enumerate(records, 1):
        company = cell(rec, "기업명")
        demand_no = cell(rec, "번호")
        self_kw = name_idx.get(norm_name(company), [])
        demand_kw = []
        for col in ("수요기술명", "수요기술 내용", "예상 적용 제품 및 서비스"):
            txt = cell(rec, col)
            if txt:
                demand_kw += uik.normalize_user_input(txt)
        combined = dedup(list(self_kw) + demand_kw)
        kw_string = SEP.join(combined)

        cos = np.clip(M @ encode(kw_string), 0, 1)
        match = W_COS * cos + W_PROM * pnorm
        rankscore = match if RANK_BY == "match" else cos
        cand_idx = np.argsort(-rankscore)[:topk]

        # 임베딩 Top10 (참고 시트)
        for rank, i in enumerate(cand_idx[:TOPN_FINAL], 1):
            emb_rows.append({
                "번호": demand_no, "기업명": company, "기술번호": cell(rec, "기술번호"),
                "수요기술명": cell(rec, "수요기술명"), "rank": rank,
                "과제고유번호": pid[i], "과제명": pname[i], "과제수행기관": org[i],
                "유사도_코사인": round(float(cos[i]), 6),
                "유망성점수": round(float(promise[i]), 4),
                "매칭점수": round(float(match[i]), 6),
                "과제설명문": pdesc[i][:DESC_OUT],
            })

        # LLM 재랭킹 (체크포인트 재사용)
        demand = {c: cell(rec, c) for c in
                  ("수요기술명", "수요기술 내용", "수요기술 사양", "예상 적용 제품 및 서비스")}
        demand["기업명"] = company
        demand["keywords"] = combined
        cands = [{"id": j + 1, "과제명": pname[i], "과제설명문": pdesc[i],
                  "키워드": list(pkw[i]) if isinstance(pkw[i], (list, tuple)) else []}
                 for j, i in enumerate(cand_idx)]
        ck = str(demand_no)
        if ck in ckpt:
            id2 = {int(k): tuple(v) for k, v in ckpt[ck].items()}
        else:
            id2 = llm_score(demand, cands)
            ckpt[ck] = {str(k): list(v) for k, v in id2.items()}
            with open(LLM_CKPT, "w", encoding="utf-8") as f:
                json.dump(ckpt, f, ensure_ascii=False)

        # id(1..K) → 코퍼스 인덱스 매핑하여 LLM 점수로 정렬
        scored = []
        for j, i in enumerate(cand_idx):
            s, reason = id2.get(j + 1, (0, ""))
            scored.append((s, j, i, reason))
        scored.sort(key=lambda x: (-x[0], -float(match[x[2]])))
        for rank, (s, j, i, reason) in enumerate(scored[:TOPN_FINAL], 1):
            row = {
                "번호": demand_no, "기업명": company, "기술번호": cell(rec, "기술번호"),
                "수요기술명": cell(rec, "수요기술명"), "rank": rank,
                "LLM점수": s, "LLM판단근거": reason,
                "과제고유번호": pid[i], "과제명": pname[i], "과제수행기관": org[i],
                "유사도_코사인": round(float(cos[i]), 6),
                "유망성점수": round(float(promise[i]), 4),
                "과제설명문": pdesc[i][:DESC_OUT],
            }
            if args.explain:
                ekey = f"{demand_no}::{pid[i]}"
                if ekey not in ex_ckpt:
                    meta = pmeta.get(pid[i], {}) if isinstance(pmeta, dict) else {}
                    proj = {
                        "pid": pid[i], "과제명": pname[i], "설명": pdesc[i],
                        "유망성": round(float(promise[i]), 1), "수행기관": org[i],
                        "키워드": list(pkw[i]) if isinstance(pkw[i], (list, tuple)) else [],
                        "논문명": meta.get("논문명_리스트") or [],
                        "특허명": meta.get("특허명_리스트") or [],
                        "논문건수": meta.get("논문건수", 0), "특허건수": meta.get("특허건수", 0),
                        "총연구비_상위비율": meta.get("총연구비_상위비율"),
                        "논문건수_상위비율": meta.get("논문건수_상위비율"),
                        "특허건수_상위비율": meta.get("특허건수_상위비율"),
                    }
                    ex_ckpt[ekey] = gen_explanation(demand, proj)[:DESC_OUT]
                row["추천근거_상세"] = ex_ckpt[ekey]
            llm_rows.append(row)
        if args.explain:
            with open(EXPLAIN_CKPT, "w", encoding="utf-8") as f:
                json.dump(ex_ckpt, f, ensure_ascii=False)
        best = scored[0][0] if scored else 0
        print(f"      [{demand_no}] {company}: 후보{len(cand_idx)} "
              f"self_kw={len(self_kw)} -> LLM best={best} "
              f"({'ckpt' if ck in ckpt and ck != str(demand_no) else 'new'}) [{n}/{len(records)}]")

    print(f"[6/6] '{OUT_XLSX}' 기록…")
    book = load_workbook(XLSX)
    for sheet, rows in ((LLM_SHEET, llm_rows), (EMB_SHEET, emb_rows)):
        if sheet in book.sheetnames:
            del book[sheet]
        ws = book.create_sheet(sheet)
        out = pd.DataFrame(rows)
        ws.append(list(out.columns))
        for r in out.itertuples(index=False):
            ws.append(list(r))
    book.save(OUT_XLSX)
    print(f"완료: '{LLM_SHEET}' {len(llm_rows)}행, '{EMB_SHEET}' {len(emb_rows)}행 → {OUT_XLSX}")


if __name__ == "__main__":
    main()
