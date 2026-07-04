# -*- coding: utf-8 -*-
"""COMPA 최종추천(보완) xlsx 를 기업별 탭(시트)으로 분리 저장한다.
- 기업명 하나당 시트 1개(같은 기업의 여러 수요·rank 행은 한 시트에 모음).
- 모든 셀 wrap_text=True + 상단정렬 → 셀 안의 줄바꿈(\\n)이 여러 줄로 보이게 한다.
- 긴 텍스트 컬럼은 넓게, 헤더는 굵게+고정.
원본은 수정하지 않고 별도 파일(_기업별.xlsx)로 저장한다.
"""
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 컬럼별 너비(글자 수 기준). 미지정 컬럼은 DEFAULT_W.
WIDTHS = {
    "번호": 6, "기업명": 16, "기업임베딩매칭": 12, "기술번호": 12,
    "수요기술명": 34, "키워드": 30, "rank": 5, "LLM점수": 7,
    "LLM판단근거": 45, "과제고유번호": 14, "과제명": 40, "과제수행기관": 22,
    "유사도_과제코사인": 12, "유사도_기업코사인": 12, "유망성점수": 9,
    "추천근거_상세": 70, "과제설명문": 60, "기업설명문_사용": 10,
    "추천근거_상세_보완": 70,
}
DEFAULT_W = 16
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
HEAD_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
HEAD_FONT = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")


def safe_sheet_name(name, used):
    """엑셀 시트명 제약(31자, []:*?/\\ 금지, 공백/중복) 처리."""
    s = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name)).strip() or "무명"
    s = s[:31]
    base, i = s, 1
    while s.lower() in used:
        suf = f"_{i}"
        s = base[:31 - len(suf)] + suf
        i += 1
    used.add(s.lower())
    return s


def split_file(src, dst):
    df = pd.read_excel(src)
    wb = Workbook()
    wb.remove(wb.active)
    cols = list(df.columns)
    used = set()
    companies = list(dict.fromkeys(df["기업명"].astype(str).tolist()))
    for comp in companies:
        sub = df[df["기업명"].astype(str) == comp]
        ws = wb.create_sheet(safe_sheet_name(comp, used))
        # 헤더
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = HEAD_FONT
            cell.alignment = HEAD_ALIGN
            cell.fill = HEAD_FILL
        # 데이터
        for r, (_, row) in enumerate(sub.iterrows(), 2):
            for c, col in enumerate(cols, 1):
                v = row[col]
                if pd.isna(v):
                    v = ""
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = WRAP
        # 너비/고정/자동필터
        for c, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(col, DEFAULT_W)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    wb.save(dst)
    print(f"  → '{dst}' 저장 (기업 탭 {len(companies)}개, 총 {len(df)}행)")


def main():
    pairs = [
        ("COMPA_김민주_최종추천_보완.xlsx", "COMPA_김민주_최종추천_보완_기업별.xlsx"),
        ("COMPA_이중연_최종추천_보완.xlsx", "COMPA_이중연_최종추천_보완_기업별.xlsx"),
    ]
    if len(sys.argv) >= 3:
        pairs = [(sys.argv[1], sys.argv[2])]
    for src, dst in pairs:
        print(f"[분리] {src}")
        split_file(src, dst)
    print("완료.")


if __name__ == "__main__":
    main()
