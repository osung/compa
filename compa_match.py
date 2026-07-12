#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""COMPA 진성수요 ↔ 국가 R&D 과제 매칭 — 단일 파일 스크립트.

원래 8개 모듈(match_compa_2stage / match_compa_list / matching_viewer_explain /
user_input_keywords / postprocess_compa / split_by_company + 두 llm_pipeline 의
프롬프트)을 하나로 합친 self-contained 버전이다. 동작은 원본 파이프라인과 동일하며,
프롬프트(SYSTEM_PROMPT / FEWSHOT_MESSAGES)는 원본에서 byte 단위로 추출·주입했다.

파이프라인 (기업 수요 1건 = 시트 한 행):
    1) 키워드 추출  : 수요기술명·내용·사양 → LLM 으로 핵심 기술 키워드 10~20개
    2) 임베딩       : pro-sroberta 로 키워드 문자열 인코딩 → L2 정규화 (query 벡터)
    3) 후보 검색    : (임베딩 top N_EMB) ∪ (BM25 top N_BM) 합집합 — 리콜 확보
    4) 재랭킹       : 크로스인코더(bge-reranker)로 관련도순 → 상위 TOPK(기본 200)
    5) LLM 적합도   : TOPK 후보를 0~100 정량 평가(배치) → 상위 FINAL(기본 5)
    6) 상세 근거    : FINAL 개에 한해 4섹션 상세 추천 근거 생성
    7) 산출/후처리  : 최종추천 xlsx/pkl + 15컬럼 정리본 + 기업별 탭본

실행:
    python compa_match.py [--all | --assignee 이름[,이름]] [--exclude 이름]
                          [--topk 100] [--final 5] [--limit N]
                          [--no-explain] [--keywords-only]

필요 파일(같은 디렉토리): COMPA_진성수요_원본.xlsx, pro-sroberta/,
    public_RnD_embeddings_pro_260601_with_desc.pkl,
    company_embeddings_pro_260514_with_desc.pkl, project_match_data_260612.pkl

LLM 백엔드: Apple Silicon → MLX, CUDA → vLLM (자동 감지).
    환경변수 MV_LLM_BACKEND / MV_MLX_MODEL / MV_VLLM_MODEL 등으로 조정.

산출물·체크포인트는 담당자별로 분리 저장되며, 재실행 시 체크포인트를 재사용해
이미 처리한 수요는 건너뛴다(중단/재개 가능).
"""
import argparse
import json
import os
import pickle
import re
import sys
import threading

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer


# --- 입출력 디렉토리 (환경변수로 지정; 미지정 시 현재 디렉토리) ------------------
# Colab 에서는 COMPA_DATA_DIR 를 Google Drive 의 데이터 폴더로, COMPA_OUT_DIR 를
# Drive 의 산출물 폴더로 지정하면 세션이 끊겨도 체크포인트에서 이어서 실행할 수 있다.
DATA_DIR = os.environ.get("COMPA_DATA_DIR", ".")   # 입력: pkl·pro-sroberta·xlsx 위치
OUT_DIR = os.environ.get("COMPA_OUT_DIR", ".")     # 산출: xlsx·pkl·체크포인트 위치
os.makedirs(OUT_DIR, exist_ok=True)


# =====================================================================
# [1] LLM 백엔드(자동 선택) + 프롬프트 + 근거 생성
# =====================================================================
# Apple Silicon(macOS arm64) → MLX-LM, CUDA GPU → vLLM. 모델 ID 는 환경변수로 조정.
MLX_MODEL_ID = os.environ.get("MV_MLX_MODEL", "mlx-community/Qwen3.5-35B-A3B-4bit")
VLLM_MODEL_ID = os.environ.get("MV_VLLM_MODEL", "Qwen/Qwen3.5-35B-A3B-GPTQ-Int4")


def _detect_backend():
    """실행 환경에 맞는 LLM 백엔드('mlx'|'vllm')를 결정. MV_LLM_BACKEND 로 강제 가능."""
    forced = os.environ.get("MV_LLM_BACKEND", "").strip().lower()
    if forced in ("mlx", "vllm"):
        return forced
    import platform
    if sys.platform == "darwin" and platform.machine() == "arm64":
        return "mlx"                        # Apple Silicon → MLX
    try:
        import torch
        if torch.cuda.is_available():
            return "vllm"                   # CUDA 사용 가능 → vLLM
    except Exception:
        pass
    return "mlx" if sys.platform == "darwin" else "vllm"


BACKEND = _detect_backend()
MODEL_ID = MLX_MODEL_ID if BACKEND == "mlx" else VLLM_MODEL_ID


# ---- 방향별 프롬프트 (빌드 시 원본 llm_pipeline.py 에서 byte-exact 추출·주입) ----
# company : 기업→과제 추천('추천 과제의 우수성')  ← company_to_project/llm_pipeline.py
# project : 과제→기업 추천('추천 기업 우수성')    ← project_to_company/llm_pipeline.py
# 본 COMPA 파이프라인은 direction="company" 만 사용하지만, 원본과 동일하게 둘 다 둔다.
SYSTEM_PROMPT_COMPANY = "너는 추천 시스템의 '추천 이유'를 설명하는 한국어 AI야. 입력 JSON의 정보를 기반으로 특정 회사에 특정 과제가 왜 추천되었는지 한국어로 설명한다. 기술적 원리나 설명이 필요한 경우에는 일반적인 산업/기술 지식을 활용해 쉽게 풀어 설명할 수 있다. 다만 입력 정보와 무관한 새로운 사실(특정 기업의 추가 사업, 특정 수치, 특정 기술 보유 등)을 만들어서는 안 된다.아래 규칙을 기반으로 최종 결과만 생성해. 중간 추론 과정은 출력하지 마.\n1) 회사 정보 해석\n   1-1) company.purpose는 유효한 값이 있는 경우에만 해석하되, company.keyword와 기술적으로 연결되는 항목만 선택적으로 사용하며, 연결성이 낮은 항목은 설명에서 제외한다.\n   1-2) company.keyword는 유효한 값이 있는 경우에만 의미 단위로 묶어 회사의 핵심 역량과 기술 요소를 도출한다.\n   1-3) company.company_patent는 유효한 값이 있는 경우에만 검토하여 회사의 기술 축, 구현 방식, 적용 가능 분야를 파악한다. 연관성 서술에 활용할 때는 회사가 '보유한 특허'임을 분명히 드러내되(예: '~ 관련 특허를 보유'), 특허 제목 전체는 그대로 나열하지 않고 기술 주제로 묶어 표현한다.\n2) 과제 정보 해석\n   2-1) project.title, project.keywords는 유효한 값이 있는 경우에만 과제의 핵심 대상과 기술 방향을 추출한다.\n   2-2) related_research.paper 또는 related_research.patent에 유효한 값이 있는 경우에만 집중하는 핵심 기술 또는 연구 방향을 파악한다.\n3) 1)과 2)의 요약 내용을 근거로 회사와 과제의 연관성을 '연관성' 섹션에서 설명한다.\n   - 회사의 사업 및 기술 역량을 바탕으로 과제의 핵심 목적과 기술이 회사의 기술, 제품, 제조 공정 또는 연구개발과 어떻게 연결되는지 설명한다.\n   - 연결성 설명에는 반드시 1문장 이상으로 과제의 핵심 기술/방법이 왜 필요한지(작동 원리, 메커니즘, 해결하려는 문제의 원인)를 일반적인 기술 지식으로 풀어서 설명한다.\n   - 기술 설명은 '조건 또는 변수 → 그 변화로 발생하는 문제 또는 결과 → 그래서 해당 기술이 필요함'의 인과 구조로 설명한다.\n   - 연결성 설명은 다음 순서를 따른다: (회사 기술이 과제 내용에서 수행하는 역할과 의미) → (해당 기술이 실제로 활용되거나 적용될 수 있는 중간 단계) → (회사 기술/사업과의 연결) \n   - 회사명은 반드시 입력 JSON의 company.name 값을 그대로 사용한다.\n   - related_research.paper(과제의 논문 성과) 또는 related_research.patent(과제의 특허 성과)에 유효한 값이 있으면, 그 성과가 드러내는 과제의 핵심 기술·연구 방향을 회사의 기술·사업 역량과 연결지어 연관성 근거로 함께 서술한다. 다만 논문·특허 제목 전체를 그대로 나열하지 말고 공통 기술 주제로 묶어 표현하며, 값이 없으면 언급하지 않는다.\n4) 기술 설명이 필요한 경우에는 일반적인 산업 공정 지식이나 기술 지식을 활용하여 설명할 수 있다.\n   다만 입력 JSON에 없는 회사 사실이나 과제 정보를 새로 만들어서는 안 된다.\n5) 과제의 재무 및 기술 유망성, 기술 성숙도를 바탕으로 '추천 과제의 우수성' 섹션에서 작성한다.\n    - 이 섹션은 '추천된 과제 자체의 우수성'만 다룬다. 기업의 인증/지정·재무지표 등 '기업 우수성'은 이 섹션에 어떠한 형태로도 작성하지 않는다(기업 정보는 '연관성' 섹션에서만 활용).\n    - project.총연구비_상위비율과 project.총연구비_group 값이 모두 있을 때만 언급한다.\n    - project.총연구비_group이 '전체'이면 '총연구비가 전체의 상위 n% 수준'으로, 그 외에는 '총연구비가 {group} 업종 내 상위 n% 수준'으로 표현한다.\n    - project.논문건수_상위비율(n)과 project.논문건수_group(분야명)이 모두 있으면, '논문 실적이 {분야명} 분야 내 상위 n% 수준'이라고 딱 한 번만 자연스럽게 서술한다. 'project.논문건수_상위비율' 같은 필드명을 출력하거나 'n% 수준'·'상위 n%'를 중복해서 두 번 표현하지 않는다. 값이 없으면 논문 상위비율 표현은 전혀 쓰지 않는다.\n    - project.특허건수_상위비율(n)과 project.특허건수_group(분야명)이 모두 있으면, '특허 실적이 {분야명} 분야 내 상위 n% 수준'이라고 딱 한 번만 자연스럽게 서술한다. 'project.특허건수_상위비율' 같은 필드명을 출력하거나 'n% 수준'·'상위 n%'를 중복해서 두 번 표현하지 않는다. 값이 없으면 특허 상위비율 표현은 전혀 쓰지 않는다.\n    - has_paper가 true이면 '본 과제 수행을 통해 논문 paper_list_count건을 확보(발표)했다'는 식으로, 과제가 산출한 성과로 서술한다('과제와 관련된 논문'이 아니라 과제 수행의 성과임을 분명히 한다).\n    - has_patent가 true이면 '본 과제 수행을 통해 특허 patent_list_count건을 확보(출원·등록)했다'는 식으로, 과제가 산출한 성과로 서술한다('과제와 관련된 특허'가 아니라 과제 수행의 성과임을 분명히 한다).\n    - related_research.paper / related_research.patent 에 제목이 있으면, 제목 전체를 나열하지 말고 그 논문·특허들이 공통적으로 어떤 기술 분야·주제의 성과인지 한 문장 정도로 간단히 종합 설명한다.\n    - has_paper가 false이면 논문 실적 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_patent가 false이면 특허 실적 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_paper와 has_patent가 모두 false이면 논문·특허 실적 및 이를 근거로 한 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - 정량 지표는 숫자를 나열하듯 쓰지 말고, 과제의 우수성을 설명하는 보조 근거로 자연스럽게 종합 서술한다.\n    - 추천 과제 우수성 섹션은 다음 순서로 작성한다.\n      1. 과제의 총연구비 수준을 기반으로 과제 자체의 우수성을 설명한다.\n      2. has_paper 또는 has_patent 중 하나 이상이 true인 경우에만 다음 내용을 기반으로 과제의 기술 성숙도를 작성한다.\n        - has_paper가 true이면 '본 과제 수행을 통해 논문 paper_list_count건을 확보'한 성과로 설명하고, related_research.paper 제목이 있으면 어떤 기술 주제의 성과인지 간단히 종합한다.\n        - has_patent가 true이면 '본 과제 수행을 통해 특허 patent_list_count건을 확보'한 성과로 설명하고, related_research.patent 제목이 있으면 어떤 기술 주제의 성과인지 간단히 종합한다.\n        - has_paper와 has_patent가 모두 false인 경우에는 논문·특허 실적 및 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n      3. 마지막으로 과제 자체의 적합성 및 추천 타당성을 종합 결론으로 작성한다(기업 우수성은 제외).\n6) 다음 내용을 바탕으로 '유사 사례' 섹션을 생성한다.\n   - '유사 사례'는 두 축으로 구성한다: (가) 기업의 보유특허·수행 과제와 추천 과제의 논문·특허 성과 사이의 기술적 연관성, (나) 추천(매칭)된 과제를 수행한 기업들과 현재 기업의 유사성. 특히 (나)는 '현재 기업과 유사한 기업들이 이미 이 과제를 수행했다'는 협업 필터링 관점의 추천 근거로 서술한다.\n   - (단락 구성·필수) '유사 사례 및 실적' 섹션은 반드시 두 단락으로 나누어 작성하고, 두 단락 사이는 빈 줄(줄바꿈 문자 두 개 '\\n\\n')로 구분한다. 첫째 단락에는 (가) 추천 과제의 논문·특허 성과와 기업 보유특허의 기술적 연관성을, 둘째 단락에는 (나) 추천 과제를 수행한 기업(과제수행기관)과 현재 기업의 유사성(및 지역적 연관성)을 작성한다. (가) 또는 (나) 중 한쪽 내용이 없으면 해당 단락은 생략하고, 작성되는 단락이 하나뿐이면 단락 구분 줄바꿈도 넣지 않는다.\n  [매칭 과제와의 유사성]\n   - company.has_patent_matching_related가 true인 경우에만, patent_matching_related에 포함된 특허 제목들과 2)에서 파악한 현재 추천 과제의 내용을 비교하여 기술적 연관성을 설명한다.\n   - 이때 해당 내용이 회사가 '보유한 특허'임을 문장에서 분명히 드러낸다(예: '~ 관련 특허를 보유하여'). 다만 특허 제목 전체를 그대로 나열하지는 말고, 특허들이 공통적으로 가리키는 기술 주제, 적용 분야, 해결하려는 문제와 현재 추천 과제의 목표·핵심 기술·적용 방식이 어떻게 연결되는지 종합하여 설명한다.\n   - company.has_patent_matching_related가 false이면 특허 매칭 관련 내용을 어떠한 형태로도 작성하지 않는다.\n [매칭 과제의 유사 논문 및 특허 성과]\n    - has_paper 또는 has_patent 중 하나 이상이 true인 경우에만 매칭 과제의 유사 논문 및 특허 성과 내용을 작성한다.\n    - has_paper와 has_patent가 모두 false이면 매칭 과제의 유사 논문 및 특허 성과 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_paper가 false이면 논문 실적, 논문 부재, 논문 기반 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_patent가 false이면 특허 실적, 특허 부재, 특허 기반 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - 논문 또는 특허가 여러 개 제공되는 경우 각 항목을 단순 나열하지 말고 반복되는 기술 주제 또는 공통 연구 방향을 중심으로 종합적으로 설명한다.\n    - has_paper 또는 has_patent 중 true인 항목의 내용만 종합하여 해당 과제가 어떤 기술 분야나 연구 방향에 집중하고 있는지 정리하고 회사 목적 및 기술과 어떻게 연관 되는지 설명한다.\n    - 이후 해당 기술이 일반적으로 어떤 장치나 시스템에서 사용되는지 설명하고, 그 장치 또는 시스템이 1)에서 파악한 현재 회사의 사업과 어떻게 연결되는지 설명한다.\n    - has_paper 또는 has_patent 중 하나 이상이 true인 경우, 과제의 기술 → 기술이 필요한 이유 → 기술이 사용되는 장치 또는 시스템 → 회사 사업과의 연관 근거 순서로 설명한다.\n   - conduct_list_company와 conduct_list_project는 각각 독립적으로 판단한다.\n   - 빈 리스트, None, 누락된 값에 대해서는 어떤 형태로도 언급하지 않는다.\n   - 빈 값을 근거로 한 추측, 보완 설명, 일반화된 설명을 생성하지 않으며, '없다', '제공되지 않았다', '비어 있다', '확인되지 않는다'와 같은 표현은 절대 생성하지 않는다.\n   - (중요) 수행 기업·지역 연관성 등 어떤 정보가 없을 때, 그 부재나 생략 사실 자체를 문장으로 서술하지 않는다. '정보가 제공되지 않아', '정보가 없어', '해당 내용은 생략한다', '생략합니다', '활용되지 않는다/않으나', '작성할 수 없다', '수행 기업이 존재하지 않아' 같은 표현이나 conduct_list_company 같은 내부 필드명은 절대 출력하지 않는다. 근거가 없는 부분은 그냥 작성하지 말고, 근거가 있는 내용만 자연스럽게 서술한 뒤 문장을 끝낸다.\n   [추천된 과제를 수행한 기업 유사 사례]\n   - has_conduct_company가 True인 경우에만 추천된 과제를 수행한 기업 유사 사례 내용을 작성한다.\n   - has_conduct_company가 False이면 수행 기업, 수행 기업 부재, 수행 기업 유사 사례 관련 내용을 어떠한 형태로도 작성하지 않는다.\n   - (중요·먼저) 이 블록의 회사들은 '추천(매칭)된 과제를 실제로 수행한 기업'이다. 이 블록은 둘째 단락의 맨 앞에 오며, 반드시 그 첫 문장을 '과제수행기관인 {company_name}은 …' 형태로 시작해 그 회사가 추천 과제를 수행한 기관임을 먼저 명시한 뒤 비교 설명을 이어간다. 회사 이름만 단독으로(예: '{company_name}은 …') 시작하거나, '이 과제와 유사한 기술을 보유한 기업', '유사한 사업을 수행하는 기업' 처럼 수행 사실을 흐리는 표현은 쓰지 않는다.\n   - conduct_list_company에서는 각 항목의 company_info 안에 있는 company_name, company_purpose_list, company_keyword, region, company_patent(수행기관 보유특허), conduct_projects(수행기관이 수행한 과제) 중 값이 존재하는 정보만 활용한다.\n   - 특히 수행기관의 company_patent·conduct_projects 와 현재(매칭) 기업의 보유특허(patent_matching_related)·수행과제(conduct_list_project)를 비교하여, 두 기업이 어떤 기술 주제의 특허·수행과제에서 공통되는지 분석한다(협업 필터링: 매칭 기업과 유사한 특허·수행과제를 가진 기업이 이미 이 과제를 수행함을 근거로 제시).\n   - 추천된 과제를 수행한 기업(수행 기업)의 company_info(보유특허·수행과제·키워드·사업목적·지역 등)와 1)에서 파악한 현재(분석/추천) 기업의 정보를 '두 기업끼리만' 비교하여, 두 기업이 어떤 기술·사업·공정·제품·연구개발 방향에서 유사한지 설명한다.\n   - (금지) 수행 기업과 '추천(매칭)된 과제' 사이의 연관성(수행 기업이 그 과제를 수행했다는 사실, 수행 기업 기술이 과제와 맞다는 식의 설명)은 자명하므로 절대 분석하거나 언급하지 않는다. 이 블록의 비교 대상은 오직 '수행 기업 ↔ 현재(분석/추천) 기업'이며, 추천 과제와의 연관성은 다른 블록에서만 다룬다.\n   - 단순히 '유사하다'고 표현하지 말고, 수행 기업의 기술 또는 사업 내용이 현재(분석/추천) 기업과 어떤 점에서 공통적으로 닮아 있는지 구체적으로 서술한다.\n   - 공통점이 분명한 경우, '현재 기업과 유사한 기업들이 이미 이 과제를 수행했다'는 점을 추천의 핵심 근거(협업 필터링 관점)로 분명히 제시한다.\n   - 단, 수행 회사의 기술·사업·공정·제품·연구개발 방향이 현재 회사 및 추천 과제와 실질적인 공통점이 뚜렷하지 않으면(연관성이 낮으면) 억지로 유사하다고 엮지 말고, 해당 수행 회사에 대한 내용을 유사 사례에 포함하지 않는다(언급 자체를 생략한다). 공통점이 분명한 수행 회사가 하나도 없으면 '추천된 과제를 수행한 기업 유사 사례' 자체를 작성하지 않는다.\n   - conduct_list_company에 여러 항목이 있을 경우 각 회사를 하나씩 단순 나열하지 말고, 반복되는 공통 기술 주제, 공통 사업 방향, 공통 문제 해결 방식 중심으로 종합하여 설명한다.\n   [추천 과제를 수행한 기업의 지역적 연관성]\n   - 지역 근거는 has_conduct_company가 True인 경우에만 보조 근거로 활용할 수 있다.\n   - has_conduct_company가 False이면 지역적 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n   - 지역 근거는 기술적 유사성의 대체가 아니라 보조 근거로만 사용하며, 같은 지역 또는 유사 지역이라는 이유만으로 기술적 적합성을 단정하지 않는다.\n   - conduct_list_company의 각 항목에 포함된 company_info.region과 company.region을 비교하여 지역적 연관성을 판단한다.\n   - 현재 회사와 동일한 지역의 수행 기업이 있으면 유사 지역보다 우선적으로 설명한다.\n   - 동일 지역 기업이 여러 개이면 company.region 값을 직접 언급하여 지역적 연관성이 비교적 강한 보조 근거임을 설명한다.\n   - 동일 지역 기업이 없고 유사 지역 기업이 있을 때만 유사 지역 근거를 사용한다.\n   - 유사 지역은 수도권(서울, 경기, 인천), 충청권(충남, 충북, 대전, 세종), 호남권(전북, 전남, 광주), 영남권(경남, 경북, 대구, 부산), 강원권(강원) 기준으로 판단한다.\n   - 강원권은 정확히 '강원'이 일치할 때만 지역 근거로 사용한다.\n   - region_relation.has_same_region이 true이면 동일 지역 근거를 보조적으로 사용하고, company.region 값을 직접 언급한다.\n   - region_relation.has_same_region이 false이고 region_relation.has_similar_region이 true이면 유사 권역 근거를 보조적으로 사용하고, region_relation.company_region_group 값을 직접 언급한다.\n   [기업이 수행한 과제 유사 사례]\n   - has_conduct_project가 True인 경우에만 기업이 수행한 과제 유사 사례 내용을 작성한다.\n   - has_conduct_project가 False이면 수행 과제, 수행 이력, 과거 과제 부재와 관련된 내용을 어떠한 형태로도 작성하지 않는다.\n   - conduct_list_project에서는 각 항목의 project_info 안에 있는 project_name, project_keyword, paper, patent 중 값이 존재하는 정보만 활용한다.\n   - conduct_list_project에 여러 항목이 있을 경우 각 과제를 하나씩 단순 나열하지 말고, 공통 기술 주제, 공통 연구개발 방향, 공통 문제 해결 방식 중심으로 종합하여 설명한다.\n   - 현재 회사가 과거에 수행했던 과제들의 project_info와 2)에서 파악한 현재 추천된 과제의 내용을 비교하여 목표, 핵심 기술, 적용 방식, 해결하려는 문제 측면에서 어떤 유사성이 있는지 설명한다.\n7) 모든 요소를 섹션 구조로 일관되게 정리한다.\n8) 모든 섹션은 일반인이 이해할 수 있는 수준으로 작성한다.\n\n[출력 규칙]\n- 위 내부 사고 단계나 중간 판단, 계산 과정은 절대 출력하지 않는다.\n- 추측하거나 없는 사실을 만들지 않는다.\n- 입력 JSON에 값이 없는 항목은 언급하지 않는다.\n- 다만 기술 설명이나 원리 설명이 필요한 경우에는 일반적인 산업 또는 기술 지식을 활용하여 설명할 수 있다.\n- 모든 설명은 '평가'가 아니라 '추천 이유 설명'의 관점에서 작성한다.\n"
FEWSHOT_MESSAGES_COMPANY = [{'role': 'user', 'content': '아래 JSON만을 근거로 추천 근거를 작성해.\n출력은 JSON 하나만 반환해. (JSON 외 텍스트 금지)\n키는 output_requirements.format에 있는 섹션 제목을 그대로 사용해.\n\n{"company": {"company_id": "C001", "name": "샘플회사_알파12a", "purpose": ["반도체 및 평판디스플레이 제조용 기계 제조업", "항공기, 우주선 및 보조장치 제조업", "공학 연구개발업"], "keyword": ["코팅공정용", "나노물질자가정렬방법", "코팅물질", "용액공정", "용액기반", "공압모듈", "코팅장비마스크패터닝", "코팅장비", "코팅솔루션", "필름제조용", "금속입자소결체", "진공증착", "공압부품", "잉크토출장치", "공정기술", "피인쇄물질", "밸브제작", "용액"], "patent_matching_related": [{"특허명": "프린팅 장치"}, {"특허명": "유도보조 전극을 포함하는 유도 전기수력학 젯 프린팅 장치"}, {"특허명": "피드백 제어형 인쇄 시스템"}, {"특허명": "전기수력학 방식의 분사 노즐"}], "has_patent_matching_related": true, "conduct_list_project": [{"과제명": "고분자 기반 기능성 필름 제조 공정 개발", "과제코드": "P9001", "project_info": {"project_name": "고분자 기반 기능성 필름 제조 공정 개발", "project_keyword": ["고분자", "필름", "코팅", "건조", "표면 제어"], "paper": [], "patent": []}}, {"과제명": "정밀 프린팅 기반 소재 패터닝 기술 개발", "과제코드": "P9002", "project_info": {"project_name": "정밀 프린팅 기반 소재 패터닝 기술 개발", "project_keyword": [], "paper": null, "patent": null}}], "has_conduct_project": true, "region": "서울", "벤처기업여부": "Y", "이노비즈여부": "Y", "메인비즈여부": "N", "ASTI 여부": "N", "특구 여부": "N", "매출성장율_상위비율": "5", "매출성장율_group": "전체", "영업이익율_상위비율": "", "영업이익율_group": null, "부채비율_하위비율": "10", "부채비율_group": "전체", "연구개발비_상위비율": "20", "연구개발비_group": "전문직별 공사업"}, "project": {"project_id": "P001", "title": "액정 엘라스토머 기반 4D 프린팅 소재 개발", "keyword": ["코팅공정용", "용액공정", "금속입자소결체", "잉크토출장치", "공정기술"], "project_score": 70.21047, "cosine_distance": 0.024936, "description": "이 과제는 테스트기관베타7X에서 수행했으며, 주조/용접/접합 분야에 속합니다. 또한, 총연구비는 전체의 상위 5%에 속하며, 총연구기간은 8개월 30일입니다. 이 과제는 액정 엘라스토머 소재를 기반으로 시간이 흐르면서 형태가 변하는 4D 프린팅 소재를 개발하는 연구입니다.이 소재는 고분자 탄성 액추에이터와 관련된 4D 프린팅 기술을 활용하며, 액정 엘라스토머 필름 제조 방법 등 기능성 소재의 제작 기술을 다룹니다.논문과 특허를 보면 형태 변화가 가능한 소재와 그 제조 기술에 대한 연구가 중심입니다.", "총연구비_상위비율": "5", "총연구비_group": "전체"}, "conduct_list_company": [{"company_name": "가상회사_유동제어_1", "company_id": "C101", "company_info": {"company_name": "가상회사_유동제어_1", "region": "서울", "company_keyword": ["기능성 필름", "정밀 코팅", "박막 형성", "공정 제어"], "company_purpose_list": ["디스플레이 및 전자재료용 소재 개발"], "company_patent": []}}, {"company_name": "임시조직a12k", "company_id": "C102", "company_info": {"company_name": "임시조직a12k", "region": "경기", "company_keyword": [], "company_purpose_list": [], "company_patent": []}}], "has_conduct_company": true, "region_relation": {"company_region": "서울", "company_region_group": "수도권", "same_region_count": 1, "similar_region_count": 1, "same_region_companies": [{"company_name": "가상회사_유동제어_1", "region": "서울"}], "similar_region_companies": [{"company_name": "임시조직a12k", "region": "경기"}], "same_region_names": ["가상회사_유동제어_1"], "similar_region_names": ["임시조직a12k"], "same_region_group": "수도권", "similar_region_group": "수도권", "has_same_region": true, "has_similar_region": true}, "related_research": {"paper": ["습윤 고분자 탄성 액추에이터의 4D 프린팅"], "patent": ["폴리로탁산 가교체를 도입한 액정 엘라스토머 필름의 제조 방법"]}, "paper_list_count": 1, "patent_list_count": 1, "has_paper": true, "has_patent": true, "output_requirements": {"language": "ko", "format": ["연관성", "추천 과제의 우수성", "유사 사례 및 실적"], "section_sentence_range": "각 섹션 4~6문장", "forbidden": ["예시", "참고", "제출", "메타"]}}'}, {'role': 'assistant', 'content': '{"연관성": "샘플회사_알파12a는 코팅공정, 용액기반 공정, 잉크 토출 장치, 공압 모듈 및 정밀 코팅 장비와 관련된 기술 요소를 보유하고 있어 액정 엘라스토머 기반 4D 프린팅 소재 개발 과제와 높은 기술적 연결성을 가진다.액정 엘라스토머 기반 4D 프린팅은 소재 내부 배열과 코팅 균일도가 변하면 형태 변화 성능과 구동 안정성이 달라질 수 있기 때문에, 용액 공정과 정밀 코팅 기술을 통해 소재의 두께와 배열을 안정적으로 제어하는 기술이 중요하다.샘플회사_알파12a의 코팅장비, 공정기술, 잉크토출장치 관련 역량은 기능성 필름 제조와 정밀 소재 패터닝 과정에서 활용될 수 있으며, 이는 액정 엘라스토머 필름 제조 및 4D 프린팅 공정의 중간 제조 단계와 직접 연결된다.특히 금속입자소결체와 용액기반 공정 기술은 기능성 소재를 균일하게 형성하고 패턴 정밀도를 높이는 데 활용될 수 있어 고분자 기반 액추에이터 제작 공정과의 연관성이 높다.또한 프린팅 장치, 피드백 제어형 인쇄 시스템, 전기수력학 방식의 분사 노즐과 같은 특허들은 정밀 분사 및 인쇄 공정 안정화와 관련된 기술 축을 형성하고 있으며, 이는 4D 프린팅 소재 제조 과정에서 필요한 정밀 패터닝 및 균일 코팅 기술과 자연스럽게 연결된다.", "추천 과제의 우수성": ["추천된 과제는 총연구비가 전체의 상위 5% 수준에 해당하여 연구개발 규모 측면에서 우수한 과제로 볼 수 있습니다.또한 관련 논문 1건과 특허 1건의 실적이 있어, 형태 변화가 가능한 고분자 소재와 필름 제조 기술에 대한 연구 기반이 확인됩니다.종합하면 이 과제는 연구개발 규모와 관련 논문·특허 성과 측면에서 기술적 완성도와 연구 기반이 탄탄하여 추천 타당성이 높은 과제입니다."], "유사 사례 및 실적": "추천된 과제는 본 과제 수행을 통해 습윤 고분자 탄성 액추에이터·액정 엘라스토머 필름 제조 관련 논문·특허를 확보했으며, 이는 형태 변화형 고분자 소재와 기능성 필름 제작 분야의 성과입니다.샘플회사_알파12a가 보유한 프린팅 장치·전기수력학 분사 노즐 등 정밀 토출·패턴 형성 관련 특허는 이러한 고분자 소재를 균일하게 제어·분사하는 기술과 맞닿아 있어, 과제의 연구성과와 기술적으로 연관됩니다.\\n\\n과제수행기관인 가상회사_유동제어_1은 기능성 필름·정밀 코팅·박막 형성·공정 제어 기술을 보유하고 있으며, 이는 현재 기업인 샘플회사_알파12a의 코팅·증착·소재 제조 역량과 직접적으로 공통됩니다.두 기업 모두 박막·코팅 공정과 소재 제어 기술을 사업 기반으로 삼고 있어, 현재 기업과 유사한 기업이 이미 이 과제를 수행했다는 점(협업 필터링)이 추천을 뒷받침합니다.또한 과제수행기관 가상회사_유동제어_1은 현재 기업과 동일한 서울 지역에 위치해 지역적 연관성도 보조 근거가 됩니다."}'}]
SYSTEM_PROMPT_PROJECT = "너는 추천 시스템의 '추천 이유'를 설명하는 한국어 AI야. 입력 JSON의 정보를 기반으로 특정 회사에 특정 과제가 왜 추천되었는지 한국어로 설명한다. 기술적 원리나 설명이 필요한 경우에는 일반적인 산업/기술 지식을 활용해 쉽게 풀어 설명할 수 있다. 다만 입력 정보와 무관한 새로운 사실(특정 기업의 추가 사업, 특정 수치, 특정 기술 보유 등)을 만들어서는 안 된다.아래 규칙을 기반으로 최종 결과만 생성해. 중간 추론 과정은 출력하지 마.\n1) 회사 정보 해석\n   1-1) company.purpose는 유효한 값이 있는 경우에만 해석하되, company.keyword와 기술적으로 연결되는 항목만 선택적으로 사용하며, 연결성이 낮은 항목은 설명에서 제외한다.\n   1-2) company.keyword는 유효한 값이 있는 경우에만 의미 단위로 묶어 회사의 핵심 역량과 기술 요소를 도출한다.\n   1-3) company.company_patent는 유효한 값이 있는 경우에만 검토하여 회사의 기술 축, 구현 방식, 적용 가능 분야를 파악한다. 연관성 서술에 활용할 때는 회사가 '보유한 특허'임을 분명히 드러내되(예: '~ 관련 특허를 보유'), 특허 제목 전체는 그대로 나열하지 않고 기술 주제로 묶어 표현한다.\n2) 과제 정보 해석\n   2-1) project.title, project.keywords는 유효한 값이 있는 경우에만 과제의 핵심 대상과 기술 방향을 추출한다.\n   2-2) related_research.paper 또는 related_research.patent에 유효한 값이 있는 경우에만 집중하는 핵심 기술 또는 연구 방향을 파악한다.\n3) 1)과 2)의 요약 내용을 근거로 과제와 회사의 연관성을 '연관성' 섹션에서 설명한다.\n   - 과제 추천 방향이므로, 먼저 추천 과제의 핵심 목적·대상·기술 방향을 설명한 뒤, 그 과제의 기술이 회사의 사업·기술·제품·제조 공정·연구개발과 어떻게 연결되는지 이어서 설명한다(과제 설명이 먼저, 기업 설명이 나중).\n   - 연결성 설명에는 반드시 1문장 이상으로 과제의 핵심 기술/방법이 왜 필요한지(작동 원리, 메커니즘, 해결하려는 문제의 원인)를 일반적인 기술 지식으로 풀어서 설명한다.\n   - 기술 설명은 '조건 또는 변수 → 그 변화로 발생하는 문제 또는 결과 → 그래서 해당 기술이 필요함'의 인과 구조로 설명한다.\n   - 연결성 설명은 다음 순서를 따른다: (추천 과제의 핵심 목적·기술과 그것이 필요한 이유) → (해당 기술이 실제로 활용되거나 적용되는 중간 단계) → (회사의 기술/사업과의 연결) \n   - 회사명은 반드시 입력 JSON의 company.name 값을 그대로 사용한다.\n   - related_research.paper(과제의 논문 성과) 또는 related_research.patent(과제의 특허 성과)에 유효한 값이 있으면, 그 성과가 드러내는 과제의 핵심 기술·연구 방향을 회사의 기술·사업 역량과 연결지어 연관성 근거로 함께 서술한다. 다만 논문·특허 제목 전체를 그대로 나열하지 말고 공통 기술 주제로 묶어 표현하며, 값이 없으면 언급하지 않는다.\n4) 기술 설명이 필요한 경우에는 일반적인 산업 공정 지식이나 기술 지식을 활용하여 설명할 수 있다.\n   다만 입력 JSON에 없는 회사 사실이나 과제 정보를 새로 만들어서는 안 된다.\n5) 기업의 재무 및 기술 역량을 바탕으로 '추천 기업 우수성' 섹션에서 작성한다.\n    - 이 섹션은 '추천된 기업 자체의 우수성'만 다룬다. 과제의 총연구비·논문·특허 등 '과제 우수성'은 이 섹션에 어떠한 형태로도 작성하지 않는다(과제 정보는 '연관성' 섹션에서만 활용).\n    - company.벤처기업여부, company.이노비즈여부, company.메인비즈여부, company.ASTI 여부, company.특구 여부 중 값이 'Y'인 항목만 선택하여 회사의 인증/지정 근거로 반영한다.\n    - company.patent_count 값이 있고 0보다 크면 '총 ○건의 특허를 보유' 처럼 기업의 기술 역량·혁신성 근거로 반영한다. company.patent_matching_related 에 특허 제목이 있으면 대표 1~3개가 공통적으로 가리키는 기술 분야로 묶어 어떤 기술의 특허를 보유했는지 간략히 덧붙인다(제목을 길게 그대로 나열하지 말 것). patent_count 가 없거나 0이면 특허 보유 관련 내용을 작성하지 않는다.\n    - company.매출성장율_상위비율과 company.매출성장율_group 값이 모두 있을 때만 언급한다.\n    - company.매출성장율_group이 '전체'이면 '매출성장율이 전체의 상위 n% 수준'으로, 그 외에는 '매출성장율이 {group} 업종 내 상위 n% 수준'으로 표현한다.\n    - company.영업이익율_상위비율과 company.영업이익율_group 값이 모두 있을 때만 언급한다.\n    - company.영업이익율_group이 '전체'이면 '영업이익율이 전체의 상위 n% 수준'으로, 그 외에는 '영업이익율이 {group} 업종 내 상위 n% 수준'으로 표현한다.\n    - company.연구개발비_상위비율과 company.연구개발비_group 값이 모두 있을 때만 언급한다.\n    - company.연구개발비_group이 '전체'이면 '연구개발비가 전체의 상위 n% 수준'으로, 그 외에는 '연구개발비가 {group} 업종 내 상위 n% 수준'으로 표현한다.\n    - company.부채비율_하위비율과 company.부채비율_group 값이 모두 있을 때만 언급한다.\n    - company.부채비율_group이 '전체'이면 '부채비율이 전체의 상위 n% 수준'으로, 그 외에는 '부채비율이 {group} 업종 내 상위 n% 수준'으로 표현한다.\n    - 위 비율 정보와 group은 해당 값이 모두 있을 때만 언급하고, 값이 없으면 완전히 생략한다.\n    - 정량 지표는 숫자를 나열하듯 쓰지 말고, 기업의 우수성을 설명하는 보조 근거로 자연스럽게 종합 서술한다.\n    - 추천 기업 우수성 섹션은 다음 순서로 작성한다.\n      1. company.* 인증/지정 항목 중 값이 'Y'인 항목이 있는 경우 회사의 기술·사업 역량 근거로 설명한다.\n      2. company.patent_count 가 0보다 크면 보유 특허 실적(총 건수 + 대표 기술 분야)을 기업의 기술 역량 근거로 서술한다.\n      3. company.* 재무지표가 존재하는 경우 회사의 성장성·수익성·재무 안정성 근거로 자연스럽게 종합 서술한다.\n      4. 마지막으로 기업 자체의 역량 및 추천 타당성을 종합 결론으로 작성한다(과제 우수성은 제외).\n6) 다음 내용을 바탕으로 '유사 사례' 섹션을 생성한다.\n   - '유사 사례'는 두 축으로 구성한다: (가) 기업의 보유특허·수행 과제와 추천 과제의 논문·특허 성과 사이의 기술적 연관성, (나) 추천(매칭)된 과제를 수행한 기업들과 현재 기업의 유사성. 특히 (나)는 '현재 기업과 유사한 기업들이 이미 이 과제를 수행했다'는 협업 필터링 관점의 추천 근거로 서술한다.\n   - (단락 구성·필수) '유사 사례 및 실적' 섹션은 반드시 두 단락으로 나누어 작성하고, 두 단락 사이는 빈 줄(줄바꿈 문자 두 개 '\\n\\n')로 구분한다. 첫째 단락에는 (가) 추천 과제의 논문·특허 성과와 기업 보유특허의 기술적 연관성을, 둘째 단락에는 (나) 추천 과제를 수행한 기업(과제수행기관)과 현재 기업의 유사성(및 지역적 연관성)을 작성한다. (가) 또는 (나) 중 한쪽 내용이 없으면 해당 단락은 생략하고, 작성되는 단락이 하나뿐이면 단락 구분 줄바꿈도 넣지 않는다.\n  [매칭 과제와의 유사성]\n   - company.has_patent_matching_related가 true인 경우에만, patent_matching_related에 포함된 특허 제목들과 2)에서 파악한 현재 추천 과제의 내용을 비교하여 기술적 연관성을 설명한다.\n   - 이때 해당 내용이 회사가 '보유한 특허'임을 문장에서 분명히 드러낸다(예: '~ 관련 특허를 보유하여'). 다만 특허 제목 전체를 그대로 나열하지는 말고, 특허들이 공통적으로 가리키는 기술 주제, 적용 분야, 해결하려는 문제와 현재 추천 과제의 목표·핵심 기술·적용 방식이 어떻게 연결되는지 종합하여 설명한다.\n   - company.has_patent_matching_related가 false이면 특허 매칭 관련 내용을 어떠한 형태로도 작성하지 않는다.\n [매칭 과제의 유사 논문 및 특허 성과]\n    - has_paper 또는 has_patent 중 하나 이상이 true인 경우에만 매칭 과제의 유사 논문 및 특허 성과 내용을 작성한다.\n    - has_paper와 has_patent가 모두 false이면 매칭 과제의 유사 논문 및 특허 성과 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_paper가 false이면 논문 실적, 논문 부재, 논문 기반 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - has_patent가 false이면 특허 실적, 특허 부재, 특허 기반 기술 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n    - 논문 또는 특허가 여러 개 제공되는 경우 각 항목을 단순 나열하지 말고 반복되는 기술 주제 또는 공통 연구 방향을 중심으로 종합적으로 설명한다.\n    - has_paper 또는 has_patent 중 true인 항목의 내용만 종합하여 해당 과제가 어떤 기술 분야나 연구 방향에 집중하고 있는지 정리하고 회사 목적 및 기술과 어떻게 연관 되는지 설명한다.\n    - 이후 해당 기술이 일반적으로 어떤 장치나 시스템에서 사용되는지 설명하고, 그 장치 또는 시스템이 1)에서 파악한 현재 회사의 사업과 어떻게 연결되는지 설명한다.\n    - has_paper 또는 has_patent 중 하나 이상이 true인 경우, 과제의 기술 → 기술이 필요한 이유 → 기술이 사용되는 장치 또는 시스템 → 회사 사업과의 연관 근거 순서로 설명한다.\n   - conduct_list_company와 conduct_list_project는 각각 독립적으로 판단한다.\n   - 빈 리스트, None, 누락된 값에 대해서는 어떤 형태로도 언급하지 않는다.\n   - 빈 값을 근거로 한 추측, 보완 설명, 일반화된 설명을 생성하지 않으며, '없다', '제공되지 않았다', '비어 있다', '확인되지 않는다'와 같은 표현은 절대 생성하지 않는다.\n   - (중요) 수행 기업·지역 연관성 등 어떤 정보가 없을 때, 그 부재나 생략 사실 자체를 문장으로 서술하지 않는다. '정보가 제공되지 않아', '정보가 없어', '해당 내용은 생략한다', '생략합니다', '활용되지 않는다/않으나', '작성할 수 없다', '수행 기업이 존재하지 않아' 같은 표현이나 conduct_list_company 같은 내부 필드명은 절대 출력하지 않는다. 근거가 없는 부분은 그냥 작성하지 말고, 근거가 있는 내용만 자연스럽게 서술한 뒤 문장을 끝낸다.\n   [추천된 과제를 수행한 기업 유사 사례]\n   - has_conduct_company가 True인 경우에만 추천된 과제를 수행한 기업 유사 사례 내용을 작성한다.\n   - has_conduct_company가 False이면 수행 기업, 수행 기업 부재, 수행 기업 유사 사례 관련 내용을 어떠한 형태로도 작성하지 않는다.\n   - (중요·먼저) 이 블록의 회사들은 '추천(매칭)된 과제를 실제로 수행한 기업'이다. 이 블록은 둘째 단락의 맨 앞에 오며, 반드시 그 첫 문장을 '과제수행기관인 {company_name}은 …' 형태로 시작해 그 회사가 추천 과제를 수행한 기관임을 먼저 명시한 뒤 비교 설명을 이어간다. 회사 이름만 단독으로(예: '{company_name}은 …') 시작하거나, '이 과제와 유사한 기술을 보유한 기업', '유사한 사업을 수행하는 기업' 처럼 수행 사실을 흐리는 표현은 쓰지 않는다.\n   - conduct_list_company에서는 각 항목의 company_info 안에 있는 company_name, company_purpose_list, company_keyword, region, company_patent(수행기관 보유특허), conduct_projects(수행기관이 수행한 과제) 중 값이 존재하는 정보만 활용한다.\n   - 특히 수행기관의 company_patent·conduct_projects 와 현재(매칭) 기업의 보유특허(patent_matching_related)·수행과제(conduct_list_project)를 비교하여, 두 기업이 어떤 기술 주제의 특허·수행과제에서 공통되는지 분석한다(협업 필터링: 매칭 기업과 유사한 특허·수행과제를 가진 기업이 이미 이 과제를 수행함을 근거로 제시).\n   - 추천된 과제를 수행한 기업(수행 기업)의 company_info(보유특허·수행과제·키워드·사업목적·지역 등)와 1)에서 파악한 현재(분석/추천) 기업의 정보를 '두 기업끼리만' 비교하여, 두 기업이 어떤 기술·사업·공정·제품·연구개발 방향에서 유사한지 설명한다.\n   - (금지) 수행 기업과 '추천(매칭)된 과제' 사이의 연관성(수행 기업이 그 과제를 수행했다는 사실, 수행 기업 기술이 과제와 맞다는 식의 설명)은 자명하므로 절대 분석하거나 언급하지 않는다. 이 블록의 비교 대상은 오직 '수행 기업 ↔ 현재(분석/추천) 기업'이며, 추천 과제와의 연관성은 다른 블록에서만 다룬다.\n   - 단순히 '유사하다'고 표현하지 말고, 수행 기업의 기술 또는 사업 내용이 현재(분석/추천) 기업과 어떤 점에서 공통적으로 닮아 있는지 구체적으로 서술한다.\n   - 공통점이 분명한 경우, '현재 기업과 유사한 기업들이 이미 이 과제를 수행했다'는 점을 추천의 핵심 근거(협업 필터링 관점)로 분명히 제시한다.\n   - 단, 수행 회사의 기술·사업·공정·제품·연구개발 방향이 현재 회사 및 추천 과제와 실질적인 공통점이 뚜렷하지 않으면(연관성이 낮으면) 억지로 유사하다고 엮지 말고, 해당 수행 회사에 대한 내용을 유사 사례에 포함하지 않는다(언급 자체를 생략한다). 공통점이 분명한 수행 회사가 하나도 없으면 '추천된 과제를 수행한 기업 유사 사례' 자체를 작성하지 않는다.\n   - conduct_list_company에 여러 항목이 있을 경우 각 회사를 하나씩 단순 나열하지 말고, 반복되는 공통 기술 주제, 공통 사업 방향, 공통 문제 해결 방식 중심으로 종합하여 설명한다.\n   [추천 과제를 수행한 기업의 지역적 연관성]\n   - 지역 근거는 has_conduct_company가 True인 경우에만 보조 근거로 활용할 수 있다.\n   - has_conduct_company가 False이면 지역적 연관성 관련 내용을 어떠한 형태로도 작성하지 않는다.\n   - 지역 근거는 기술적 유사성의 대체가 아니라 보조 근거로만 사용하며, 같은 지역 또는 유사 지역이라는 이유만으로 기술적 적합성을 단정하지 않는다.\n   - conduct_list_company의 각 항목에 포함된 company_info.region과 company.region을 비교하여 지역적 연관성을 판단한다.\n   - 현재 회사와 동일한 지역의 수행 기업이 있으면 유사 지역보다 우선적으로 설명한다.\n   - 동일 지역 기업이 여러 개이면 company.region 값을 직접 언급하여 지역적 연관성이 비교적 강한 보조 근거임을 설명한다.\n   - 동일 지역 기업이 없고 유사 지역 기업이 있을 때만 유사 지역 근거를 사용한다.\n   - 유사 지역은 수도권(서울, 경기, 인천), 충청권(충남, 충북, 대전, 세종), 호남권(전북, 전남, 광주), 영남권(경남, 경북, 대구, 부산), 강원권(강원) 기준으로 판단한다.\n   - 강원권은 정확히 '강원'이 일치할 때만 지역 근거로 사용한다.\n   - region_relation.has_same_region이 true이면 동일 지역 근거를 보조적으로 사용하고, company.region 값을 직접 언급한다.\n   - region_relation.has_same_region이 false이고 region_relation.has_similar_region이 true이면 유사 권역 근거를 보조적으로 사용하고, region_relation.company_region_group 값을 직접 언급한다.\n   [기업이 수행한 과제 유사 사례]\n   - has_conduct_project가 True인 경우에만 기업이 수행한 과제 유사 사례 내용을 작성한다.\n   - has_conduct_project가 False이면 수행 과제, 수행 이력, 과거 과제 부재와 관련된 내용을 어떠한 형태로도 작성하지 않는다.\n   - conduct_list_project에서는 각 항목의 project_info 안에 있는 project_name, project_keyword, paper, patent 중 값이 존재하는 정보만 활용한다.\n   - conduct_list_project에 여러 항목이 있을 경우 각 과제를 하나씩 단순 나열하지 말고, 공통 기술 주제, 공통 연구개발 방향, 공통 문제 해결 방식 중심으로 종합하여 설명한다.\n   - 현재 회사가 과거에 수행했던 과제들의 project_info와 2)에서 파악한 현재 추천된 과제의 내용을 비교하여 목표, 핵심 기술, 적용 방식, 해결하려는 문제 측면에서 어떤 유사성이 있는지 설명한다.\n7) 모든 요소를 섹션 구조로 일관되게 정리한다.\n8) 모든 섹션은 일반인이 이해할 수 있는 수준으로 작성한다.\n\n[출력 규칙]\n- 위 내부 사고 단계나 중간 판단, 계산 과정은 절대 출력하지 않는다.\n- 추측하거나 없는 사실을 만들지 않는다.\n- 입력 JSON에 값이 없는 항목은 언급하지 않는다.\n- 다만 기술 설명이나 원리 설명이 필요한 경우에는 일반적인 산업 또는 기술 지식을 활용하여 설명할 수 있다.\n- 모든 설명은 '평가'가 아니라 '추천 이유 설명'의 관점에서 작성한다.\n"
FEWSHOT_MESSAGES_PROJECT = [{'role': 'user', 'content': '아래 JSON만을 근거로 추천 근거를 작성해.\n출력은 JSON 하나만 반환해. (JSON 외 텍스트 금지)\n키는 output_requirements.format에 있는 섹션 제목을 그대로 사용해.\n\n{"company": {"company_id": "C001", "name": "샘플회사_알파12a", "purpose": ["반도체 및 평판디스플레이 제조용 기계 제조업", "항공기, 우주선 및 보조장치 제조업", "공학 연구개발업"], "keyword": ["코팅공정용", "나노물질자가정렬방법", "코팅물질", "용액공정", "용액기반", "공압모듈", "코팅장비마스크패터닝", "코팅장비", "코팅솔루션", "필름제조용", "금속입자소결체", "진공증착", "공압부품", "잉크토출장치", "공정기술", "피인쇄물질", "밸브제작", "용액"], "patent_matching_related": [{"특허명": "프린팅 장치"}, {"특허명": "유도보조 전극을 포함하는 유도 전기수력학 젯 프린팅 장치"}, {"특허명": "피드백 제어형 인쇄 시스템"}, {"특허명": "전기수력학 방식의 분사 노즐"}], "has_patent_matching_related": true, "patent_count": 24, "conduct_list_project": [{"과제명": "고분자 기반 기능성 필름 제조 공정 개발", "과제코드": "P9001", "project_info": {"project_name": "고분자 기반 기능성 필름 제조 공정 개발", "project_keyword": ["고분자", "필름", "코팅", "건조", "표면 제어"], "paper": [], "patent": []}}, {"과제명": "정밀 프린팅 기반 소재 패터닝 기술 개발", "과제코드": "P9002", "project_info": {"project_name": "정밀 프린팅 기반 소재 패터닝 기술 개발", "project_keyword": [], "paper": null, "patent": null}}], "has_conduct_project": true, "region": "서울", "벤처기업여부": "Y", "이노비즈여부": "Y", "메인비즈여부": "N", "ASTI 여부": "N", "특구 여부": "N", "매출성장율_상위비율": "5", "매출성장율_group": "전체", "영업이익율_상위비율": "", "영업이익율_group": null, "부채비율_하위비율": "10", "부채비율_group": "전체", "연구개발비_상위비율": "20", "연구개발비_group": "전문직별 공사업"}, "project": {"project_id": "P001", "title": "액정 엘라스토머 기반 4D 프린팅 소재 개발", "keyword": ["코팅공정용", "용액공정", "금속입자소결체", "잉크토출장치", "공정기술"], "project_score": 70.21047, "cosine_distance": 0.024936, "description": "이 과제는 테스트기관베타7X에서 수행했으며, 주조/용접/접합 분야에 속합니다. 또한, 총연구비는 전체의 상위 5%에 속하며, 총연구기간은 8개월 30일입니다. 이 과제는 액정 엘라스토머 소재를 기반으로 시간이 흐르면서 형태가 변하는 4D 프린팅 소재를 개발하는 연구입니다.이 소재는 고분자 탄성 액추에이터와 관련된 4D 프린팅 기술을 활용하며, 액정 엘라스토머 필름 제조 방법 등 기능성 소재의 제작 기술을 다룹니다.논문과 특허를 보면 형태 변화가 가능한 소재와 그 제조 기술에 대한 연구가 중심입니다.", "총연구비_상위비율": "5", "총연구비_group": "전체"}, "conduct_list_company": [{"company_name": "가상회사_유동제어_1", "company_id": "C101", "company_info": {"company_name": "가상회사_유동제어_1", "region": "서울", "company_keyword": ["기능성 필름", "정밀 코팅", "박막 형성", "공정 제어"], "company_purpose_list": ["디스플레이 및 전자재료용 소재 개발"], "company_patent": []}}, {"company_name": "임시조직a12k", "company_id": "C102", "company_info": {"company_name": "임시조직a12k", "region": "경기", "company_keyword": [], "company_purpose_list": [], "company_patent": []}}], "has_conduct_company": true, "region_relation": {"company_region": "서울", "company_region_group": "수도권", "same_region_count": 1, "similar_region_count": 1, "same_region_companies": [{"company_name": "가상회사_유동제어_1", "region": "서울"}], "similar_region_companies": [{"company_name": "임시조직a12k", "region": "경기"}], "same_region_names": ["가상회사_유동제어_1"], "similar_region_names": ["임시조직a12k"], "same_region_group": "수도권", "similar_region_group": "수도권", "has_same_region": true, "has_similar_region": true}, "related_research": {"paper": ["습윤 고분자 탄성 액추에이터의 4D 프린팅"], "patent": ["폴리로탁산 가교체를 도입한 액정 엘라스토머 필름의 제조 방법"]}, "paper_list_count": 1, "patent_list_count": 1, "has_paper": true, "has_patent": true, "output_requirements": {"language": "ko", "format": ["연관성", "추천 기업 우수성", "유사 사례 및 실적"], "section_sentence_range": "각 섹션 4~6문장", "forbidden": ["예시", "참고", "제출", "메타"]}}'}, {'role': 'assistant', 'content': '{"연관성": "샘플회사_알파12a는 코팅공정, 용액기반 공정, 잉크 토출 장치, 공압 모듈 및 정밀 코팅 장비와 관련된 기술 요소를 보유하고 있어 액정 엘라스토머 기반 4D 프린팅 소재 개발 과제와 높은 기술적 연결성을 가진다.액정 엘라스토머 기반 4D 프린팅은 소재 내부 배열과 코팅 균일도가 변하면 형태 변화 성능과 구동 안정성이 달라질 수 있기 때문에, 용액 공정과 정밀 코팅 기술을 통해 소재의 두께와 배열을 안정적으로 제어하는 기술이 중요하다.샘플회사_알파12a의 코팅장비, 공정기술, 잉크토출장치 관련 역량은 기능성 필름 제조와 정밀 소재 패터닝 과정에서 활용될 수 있으며, 이는 액정 엘라스토머 필름 제조 및 4D 프린팅 공정의 중간 제조 단계와 직접 연결된다.특히 금속입자소결체와 용액기반 공정 기술은 기능성 소재를 균일하게 형성하고 패턴 정밀도를 높이는 데 활용될 수 있어 고분자 기반 액추에이터 제작 공정과의 연관성이 높다.또한 프린팅 장치, 피드백 제어형 인쇄 시스템, 전기수력학 방식의 분사 노즐과 같은 특허들은 정밀 분사 및 인쇄 공정 안정화와 관련된 기술 축을 형성하고 있으며, 이는 4D 프린팅 소재 제조 과정에서 필요한 정밀 패터닝 및 균일 코팅 기술과 자연스럽게 연결된다.", "추천 기업 우수성": ["샘플회사_알파12a는 벤처기업과 이노비즈 인증을 보유하고 있어 기술 기반 사업화 역량을 갖춘 기업으로 볼 수 있습니다.또한 총 24건의 특허를 보유하고 있으며, 정밀 분사·전기수력학 프린팅 등 토출·인쇄 공정 제어 분야의 특허를 중심으로 기술 역량을 확보하고 있습니다.재무적으로도 매출성장율이 전체의 상위 5% 수준이고, 부채비율이 전체의 상위 10% 수준이며, 연구개발비가 전문직별 공사업 업종 내 상위 20% 수준으로 나타나 성장성·재무 안정성·연구개발 투입 측면에서 우수합니다.종합하면 이 기업은 인증과 특허 기반의 기술 역량에 견조한 재무 지표를 함께 갖추고 있어 추천 타당성이 높은 기업입니다."], "유사 사례 및 실적": "추천된 과제는 본 과제 수행을 통해 습윤 고분자 탄성 액추에이터·액정 엘라스토머 필름 제조 관련 논문·특허를 확보했으며, 이는 형태 변화형 고분자 소재와 기능성 필름 제작 분야의 성과입니다.샘플회사_알파12a가 보유한 프린팅 장치·전기수력학 분사 노즐 등 정밀 토출·패턴 형성 관련 특허는 이러한 고분자 소재를 균일하게 제어·분사하는 기술과 맞닿아 있어, 과제의 연구성과와 기술적으로 연관됩니다.\\n\\n과제수행기관인 가상회사_유동제어_1은 기능성 필름·정밀 코팅·박막 형성·공정 제어 기술을 보유하고 있으며, 이는 현재 기업인 샘플회사_알파12a의 코팅·증착·소재 제조 역량과 직접적으로 공통됩니다.두 기업 모두 박막·코팅 공정과 소재 제어 기술을 사업 기반으로 삼고 있어, 현재 기업과 유사한 기업이 이미 이 과제를 수행했다는 점(협업 필터링)이 추천을 뒷받침합니다.또한 과제수행기관 가상회사_유동제어_1은 현재 기업과 동일한 서울 지역에 위치해 지역적 연관성도 보조 근거가 됩니다."}'}]

_DIRECTIONS = {
    "company": {
        "system": SYSTEM_PROMPT_COMPANY,
        "fewshot": FEWSHOT_MESSAGES_COMPANY,
        "fmt": ["연관성", "추천 과제의 우수성", "유사 사례 및 실적"],
    },
    "project": {
        "system": SYSTEM_PROMPT_PROJECT,
        "fewshot": FEWSHOT_MESSAGES_PROJECT,
        "fmt": ["연관성", "추천 기업 우수성", "유사 사례 및 실적"],
    },
}


def _resolve_direction(direction):
    return direction if direction in _DIRECTIONS else "company"


# -------- LLM 모델 캐시 (프로세스당 1회 로드, 첫 생성 요청 때 lazy 로드) --------
_model = None
_tokenizer = None
_model_lock = threading.Lock()
_model_load_error = None


def _load_model_mlx(progress_cb=None):
    if progress_cb:
        progress_cb(f"MLX 모델 로드 중 ({MODEL_ID})...")
    from mlx_lm import load as mlx_load
    return mlx_load(MODEL_ID)


def _load_model_vllm(progress_cb=None):
    if progress_cb:
        progress_cb(f"vLLM 모델 로드 중 ({MODEL_ID})...")
    # JSON 스키마 강제(StructuredOutputsParams)가 검증된 조합과 맞도록 V0 엔진+spawn 고정.
    os.environ.setdefault("VLLM_USE_V1", "0")
    os.environ.setdefault("VLLM_WORKER_MULTIPROC_METHOD", "spawn")
    from vllm import LLM
    from transformers import AutoTokenizer
    # 메모리/분산 파라미터는 환경변수로 조정 가능.
    gpu_util = float(os.environ.get("MV_VLLM_GPU_UTIL", "0.90"))
    max_len = int(os.environ.get("MV_VLLM_MAX_LEN", "8192"))
    dtype = os.environ.get("MV_VLLM_DTYPE", "auto")
    tp_size = int(os.environ.get("MV_VLLM_TP", "1"))       # 멀티 GPU tensor parallel
    llm = LLM(
        model=MODEL_ID, trust_remote_code=True, dtype=dtype,
        gpu_memory_utilization=gpu_util, max_model_len=max_len,
        max_num_seqs=int(os.environ.get("MV_VLLM_MAX_SEQS", "1")),
        tensor_parallel_size=tp_size,
        enforce_eager=True,                                # CUDA 그래프 캡처 메모리 절약
    )
    tokenizer = AutoTokenizer.from_pretrained(MODEL_ID, trust_remote_code=True)
    return llm, tokenizer


def load_model_blocking(progress_cb=None):
    """LLM 모델을 동기 로드(첫 호출 시 디스크→메모리). BACKEND 에 맞는 백엔드 사용."""
    global _model, _tokenizer, _model_load_error
    with _model_lock:
        if _model is not None:
            return _model, _tokenizer
        if _model_load_error is not None:
            raise _model_load_error
        try:
            loader = _load_model_mlx if BACKEND == "mlx" else _load_model_vllm
            _model, _tokenizer = loader(progress_cb)
            if progress_cb:
                progress_cb(f"{BACKEND.upper()} 모델 로드 완료")
            return _model, _tokenizer
        except Exception as e:
            _model_load_error = e
            raise


def build_messages(payload, direction="company"):
    """payload(JSON) 를 근거로 추천 근거를 생성하도록 system+few-shot+user 메시지 구성."""
    d = _DIRECTIONS[_resolve_direction(direction)]
    user_prompt = (
        "아래 JSON만을 근거로 추천 근거를 작성해.\n"
        "다만 기술 설명이나 원리 설명이 필요한 경우에는 일반적인 산업 또는 기술 지식을 활용하여 설명할 수 있다.\n"
        "company.description(기업 설명문)과 project.description(과제 설명문)은 각각 기업과 과제를 요약한 "
        "핵심 컨텍스트다. 키워드(keyword)만이 아니라 이 설명문을 반드시 읽고, 기업의 사업 영역·강점과 과제의 "
        "목표·내용을 연결지어 '연관성' 섹션의 근거로 적극 활용해.\n"
        "단, company.desc_ok 가 0 이면 기업 설명문에 품질 이슈(company.desc_issue)가 있다는 뜻이므로 "
        "그 내용은 보조 참고로만 쓰고 keyword 등 다른 정보를 우선해. desc_ok 가 1 이면 설명문을 신뢰해서 사용해도 된다.\n"
        "설명문 문장을 그대로 복사하지 말고, 추천 맥락에 맞게 재구성해서 서술해.\n"
        "few-shot 예시에 나온 모든 고유명사는 예시 전용이며, 현재 출력에 절대 재사용하지 마.\n"
        "출력은 JSON 하나만 반환해. JSON 외 텍스트 금지야.\n"
        "첫 글자는 {, 마지막 글자는 } 로 끝내.\n"
        "``` 같은 코드블록은 절대 쓰지 마.\n"
        "키는 output_requirements.format에 있는 섹션 제목을 그대로 사용해.\n\n"
        f"{json.dumps(payload, ensure_ascii=False)}"
    )
    return (
        [{"role": "system", "content": d["system"]}]
        + d["fewshot"]
        + [{"role": "user", "content": user_prompt}]
    )


def _strip_think(text):
    """모델이 가끔 출력하는 <think>...</think> 블록 제거."""
    text = re.sub(r"<think>.*?</think>\s*", "", text, flags=re.DOTALL)
    text = re.sub(r"^\s*</think>\s*", "", text)
    return text.strip()


def _build_prompt(tokenizer, messages):
    """Qwen3 instruct 채팅 템플릿 적용. thinking 모드는 가능하면 끈다."""
    try:
        return tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True, enable_thinking=False)
    except TypeError:
        # 구버전 템플릿: enable_thinking 미지원
        return tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True)


def _stream_explanation_mlx(
    messages, *, max_tokens, temperature, top_p, on_token, should_stop, expected_keys=None
):
    # MLX-LM 은 JSON 스키마 강제(grammar)를 지원하지 않으므로 expected_keys 는 생성 단계에서
    # 쓰지 않는다. temperature=0.0(결정적)+프롬프트+parse_sections 의 키 강제로 구조를 보장.
    model, tokenizer = load_model_blocking()
    from mlx_lm.generate import stream_generate
    from mlx_lm.sample_utils import make_sampler
    prompt = _build_prompt(tokenizer, messages)
    sampler = make_sampler(temp=temperature, top_p=top_p)
    pieces = []
    for resp in stream_generate(model, tokenizer, prompt=prompt,
                                max_tokens=max_tokens, sampler=sampler):
        chunk = resp.text or ""
        pieces.append(chunk)
        if on_token is not None and chunk:
            on_token(chunk)
        if should_stop is not None and should_stop():
            break
    return _strip_think("".join(pieces))


def _json_object_schema(expected_keys):
    """섹션 키들을 'string 값 필수 + 추가키 불가' JSON object 스키마로 구성."""
    keys = list(expected_keys)
    return {
        "type": "object",
        "properties": {k: {"type": "string"} for k in keys},
        "required": keys,
        "additionalProperties": False,
    }


def _vllm_structured_kwargs(expected_keys):
    """vLLM SamplingParams 용 JSON 스키마 강제 인자를 버전별로 구성(미지원 시 {} → 자유 생성)."""
    if not expected_keys:
        return {}
    schema = _json_object_schema(expected_keys)
    try:                                    # 신버전
        from vllm.sampling_params import StructuredOutputsParams
        return {"structured_outputs": StructuredOutputsParams(json=schema)}
    except Exception:
        pass
    try:                                    # 구버전
        from vllm.sampling_params import GuidedDecodingParams
        return {"guided_decoding": GuidedDecodingParams(json=schema)}
    except Exception:
        pass
    return {}


def _stream_explanation_vllm(
    messages, *, max_tokens, temperature, top_p, on_token, should_stop, expected_keys=None
):
    # vLLM 오프라인 generate 는 배치(비스트리밍) → 생성 완료 후 전체 텍스트를 한 번에 콜백.
    # expected_keys 가 있으면 JSON 스키마를 강제해 항상 유효 JSON(필수 키만)을 보장.
    model, tokenizer = load_model_blocking()
    from vllm import SamplingParams
    prompt = _build_prompt(tokenizer, messages)
    params = SamplingParams(
        temperature=temperature, top_p=top_p, max_tokens=max_tokens,
        stop=["<|im_end|>"], **_vllm_structured_kwargs(expected_keys))
    outputs = model.generate([prompt], params)
    text = outputs[0].outputs[0].text or ""
    if on_token is not None and text:
        on_token(text)
    return _strip_think(text)


def stream_explanation(
    messages, *, max_tokens=2048, temperature=0.0, top_p=1.0,
    on_token=None, should_stop=None, expected_keys=None,
):
    """추천 근거/응답 생성. 백엔드에 맞는 구현으로 위임하고 전체 텍스트를 반환한다.

    temperature 기본 0.0(결정적). expected_keys(=섹션 제목 목록)를 주면 출력 구조를 보장.
    """
    fn = _stream_explanation_mlx if BACKEND == "mlx" else _stream_explanation_vllm
    return fn(messages, max_tokens=max_tokens, temperature=temperature, top_p=top_p,
              on_token=on_token, should_stop=should_stop, expected_keys=expected_keys)


# 수행 기업·지역 연관성 등 '정보 부재/생략' 자체를 서술하는 문장 제거용 패턴.
_ABSENCE_PAT = re.compile(
    r"(제공되지\s*않|정보[가는은]\s*없|해당\s*(내용|정보|사례)\S*\s*생략|"
    r"생략(합니다|한다|됩니다|됨|하)|분석할\s*수\s*없|작성할\s*수\s*없|활용되지\s*않|"
    r"수행\s*기업[이가]?\s*(존재하지\s*않|없)|존재하지\s*않|확인되지\s*않|conduct_list_company)"
)


def _scrub_absence(text):
    """'정보가 없어 생략한다' 류의 부재·생략 메타 서술 문장을 제거(한국어 종결 '다./요.' 기준 분할)."""
    if not text or not _ABSENCE_PAT.search(text):
        return text
    sents = re.split(r"(?<=[다요]\.)", text)
    kept = "".join(s for s in sents if not _ABSENCE_PAT.search(s)).strip()
    return kept if kept else text


def _section_to_text(v):
    """섹션 값을 깔끔한 문자열로 정규화.

    모델이 값을 문장 리스트로 반환하면 "['..','..']" 꼴이 되므로 리스트는 이어붙이고,
    통째로 대괄호로 감싼 문자열은 (JSON 배열이면 파싱해) 풀어준다. 끝으로 부재 서술 제거.
    """
    if isinstance(v, (list, tuple)):
        return _scrub_absence(" ".join(_section_to_text(x) for x in v).strip())
    s = str(v).strip()
    if len(s) >= 2 and s[0] == "[" and s[-1] == "]":
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return _scrub_absence(" ".join(_section_to_text(x) for x in arr).strip())
        except Exception:
            pass
        s = s[1:-1].strip()                 # 파싱 실패 시 양끝 대괄호만 제거
    return _scrub_absence(s)


def parse_sections(text, fmt=("연관성", "추천 과제의 우수성", "유사 사례 및 실적")):
    """모델이 반환한 JSON(또는 유사) 문자열에서 섹션별 텍스트를 추출(실패 시 헤더 정규식 fallback)."""
    text = text.strip()
    # 흔한 케이스: 첫 '{' ~ 마지막 '}' 구간을 JSON 으로 시도
    try:
        first, last = text.find("{"), text.rfind("}")
        if first >= 0 and last > first:
            obj = json.loads(text[first:last + 1])
            if isinstance(obj, dict):
                return {k: _section_to_text(obj.get(k, "")) for k in fmt}
    except Exception:
        pass
    # Fallback: 섹션 헤더 기반 분할 → parts: ['', sec1, body1, sec2, body2, ...]
    out = {k: "" for k in fmt}
    pattern = "|".join(re.escape(k) for k in fmt)
    parts = re.split(rf"\b({pattern})\b\s*[:：]?\s*", text)
    for i in range(1, len(parts), 2):
        sec = parts[i]
        body = parts[i + 1] if i + 1 < len(parts) else ""
        if sec in out:
            out[sec] = _section_to_text(body)
    return out


# =====================================================================
# [2] 사용자 자유입력 → 명사 키워드 정제 (LLM 키워드 추출 실패 시 폴백)
# =====================================================================
# 저장 임베딩은 모두 '짧은 명사 키워드' 문체다. 자유입력(업종 문장체 "반도체 제품 도매업")을
# 그대로 임베딩하면 행정/정책 문체로 끌려가므로, 핵심 명사만 추출해 문체를 맞춘다.
USER_INPUT_STOPWORDS = {
    "기타", "제조업", "제품", "사업", "회사", "업체", "관련", "서비스",
    "제조", "판매", "생산", "개발", "기술", "산업", "업무", "운영",
    "관리", "지원", "제공", "이용", "사용", "활용", "처리", "수행",
    "분야", "부문", "종류", "형태", "방식", "방법", "과정", "절차",
    "대상", "범위", "내용", "항목", "종목", "품목", "물품", "물자",
    "시설", "설비", "장비", "기기", "기계", "장치", "도구", "용품",
    "자재", "재료", "원료", "부품", "소재", "물질", "성분", "요소",
    "구조", "형식", "유형", "종별", "구분", "분류", "목록",
    "도매업", "부대", "일체", "일반", "임대업", "판매업", "상기",
    "각호", "공급", "호에", "서비스업", "소매업", "부동산", "제작",
    "매매", "전문", "응용", "목적", "대행업", "신품", "공업",
    "소매", "가공", "상거래", "도매", "형성", "경영", "유지",
    "기자재", "단계", "용역", "특수", "작업", "유사", "조립", "대행",
    "가공업", "무역업", "위", "각항", "부대되는", "사업일체",
    "실시", "진행", "추진", "시행", "완료", "종료", "시작", "착수",
    "설립", "설치", "구축", "도입", "적용", "반영", "포함", "제외",
    "및", "외", "내", "중", "상", "하", "전", "후", "등", "것", "수",
}
# 업종 접미사 — 제거 후 남는 핵심 명사만 사용. 예) "반도체도매업"→"반도체"
USER_INPUT_SUFFIXES = (
    "판매업", "도매업", "소매업", "서비스업", "제조업", "생산업",
    "가공업", "임대업", "대행업", "무역업", "유통업", "공업", "업",
)
# 어절 끝 조사 — 제거 후 남는 명사만 사용. 명사 말음과 충돌 잦은 1글자 조사는 의도적 제외.
USER_INPUT_JOSA = (
    "으로서", "으로써", "에서의", "에게서", "에서", "에게", "으로",
    "와의", "과의", "에의", "에", "의", "을", "를",
)
# 연결어미·용언 활용·접속어 — 자유입력 문장에 섞여 들어온 비명사 어절.
USER_INPUT_CONNECTIVES = {
    "관련된", "관련되", "위한", "통한", "대한", "관한", "따른", "인한",
    "위해", "통해", "위하여", "하는", "되는", "있는", "없는", "같은",
    "하고", "되고", "하며", "되며", "해서", "하여", "되어", "하면", "되면",
    "또는", "혹은", "그리고", "하지만", "그러나", "따라서", "그래서",
    "해당하는", "필요한", "가능한", "부대하는", "속하는", "부대되는",
}
# 불용어 동사의 활용형 어미 — '{불용어}+어미'(이용한/활용하여/개발하는)를 통째로 버리기 위함.
USER_INPUT_VERB_EOMI = ("하여", "하는", "되어", "되는", "한", "해", "된", "되", "함", "됨")


def is_meaningless_code(token):
    """과기표준분류 코드('EH030301')·순수 숫자처럼 의미 없는 토큰인지 판정(한글 포함 시 항상 보존).

    - 순수 숫자('030301')는 길이 무관 제거
    - 영문+숫자 혼합은 5글자 이상만 코드로 보고 제거(4글자 이하 '5G','MP3' 등은 보존)
    - 순수 영문 약어('AI','IoT')는 숫자가 없어 항상 보존
    """
    if re.search(r'[가-힣]', token):
        return False
    if token.isdigit():
        return True
    if len(token) >= 5 and re.fullmatch(r'[A-Za-z0-9]+', token) \
            and re.search(r'[A-Za-z]', token) and re.search(r'\d', token):
        return True
    return False


def normalize_user_input(text):
    """자유입력을 '명사 키워드' 문체로 정제 → 토큰 리스트(등장 순서 유지).

    절차: 구두점 제거·어절 분리 → 짧은/불용어/연결어미 제거 → 무의미 코드 제거 →
          업종 접미사 제거 → 어절 끝 조사 제거 → 불용어 동사 활용형 제거 → 재검사.
    """
    if not text:
        return []
    text = re.sub(r'[^\w\s가-힣a-zA-Z0-9]', ' ', str(text))
    out = []
    for w in text.split():
        wc = w.casefold()
        if len(w) < 2 or wc in USER_INPUT_STOPWORDS or wc in USER_INPUT_CONNECTIVES:
            continue
        if is_meaningless_code(w):
            continue
        stem = w
        for suf in USER_INPUT_SUFFIXES:                 # 업종 접미사 제거
            if stem.endswith(suf) and len(stem) > len(suf):
                stem = stem[:-len(suf)]
                break
        for josa in USER_INPUT_JOSA:                    # 어절 끝 조사 제거(어간 2글자 이상 유지 시)
            if stem.endswith(josa) and len(stem) - len(josa) >= 2:
                stem = stem[:-len(josa)]
                break
        # 불용어 동사 활용형 제거: 어미를 뗀 어간이 불용어이면 토큰 전체를 버린다.
        if any(stem.endswith(em) and stem[:-len(em)].casefold() in USER_INPUT_STOPWORDS
               for em in USER_INPUT_VERB_EOMI):
            continue
        stem = stem.strip()
        sc = stem.casefold()
        if len(stem) < 2 or sc in USER_INPUT_STOPWORDS or sc in USER_INPUT_CONNECTIVES:
            continue
        out.append(stem)
    return out


# =====================================================================
# [3] LLM 적합도 점수(0~100) + 4섹션 상세 추천 근거
# =====================================================================
LLM_BATCH = 20         # LLM 한 번에 평가할 과제 수
LLM_MAXTOK = 3500      # 적합도 배치당 생성 토큰 상한
DESC_CAND = 280        # LLM 프롬프트용 과제설명문 절단 길이
DESC_OUT = 30000       # 엑셀 셀 길이 한계(32767) 회피용 출력 설명문 절단
SEP = ";"              # 키워드 join 구분자


def norm_name(s):
    """기업명 정규화 — 법인격 표기/공백 제거 후 소문자화."""
    s = re.sub(r"\(주\)|\(株\)|㈜|주식회사|\(유\)|유한회사|\(재\)|재단법인|"
               r"\(사\)|사단법인|농업회사법인|\(농\)", "", str(s))
    return re.sub(r"\s+", "", s).strip().lower()


def cell(rec, col):
    """레코드에서 컬럼 값을 문자열로(결측은 '')."""
    v = rec.get(col)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


# ---- 적합도 평가(0~100) ----
_SYS = ("당신은 기업의 기술 수요와 국가 R&D 과제의 적합성을 평가하는 기술이전 전문가입니다. "
        "각 과제가 기업의 수요기술을 해결·지원하는 데 실제로 얼마나 적합한지 냉정하게 평가합니다.")
_GUIDE = (
    "각 과제가 위 [기업 기술 수요]를 해결·지원하는 데 얼마나 적합한지 0~100 정수로 평가하세요.\n"
    "  90-100: 동일 기술/제품을 직접 해결하는 매우 높은 적합\n"
    "  70-89 : 핵심 기술·응용 분야가 상당 부분 부합\n"
    "  40-69 : 일부 요소만 관련(부분 적합)\n"
    "  10-39 : 분야만 유사하거나 약하게 관련\n"
    "  0-9   : 사실상 무관\n"
    "키워드 표면 일치가 아니라 기술 내용·적용 제품 관점의 실질 적합도를 보세요.\n"
    "반드시 아래 JSON 형식으로만, 모든 과제에 대해 답하세요(reason 은 40자 이내 한국어):\n"
    '{"results": [{"id": <과제번호>, "score": <0-100 정수>, "reason": "<근거>"}, ...]}'
)


def _demand_block(d):
    return ("[기업 기술 수요]\n"
            f"- 수요기술명: {d.get('수요기술명','')}\n"
            f"- 수요기술 내용: {d.get('수요기술 내용','')}\n"
            f"- 수요기술 사양: {d.get('수요기술 사양','')}\n"
            f"- 예상 적용 제품 및 서비스: {d.get('예상 적용 제품 및 서비스','')}")


def _cand_block(cands):
    lines = ["[평가 대상 R&D 과제 목록]"]
    for c in cands:
        kw = ", ".join(c["키워드"][:8])
        lines.append(f"{c['id']}. 과제명: {c['과제명']}\n"
                     f"   키워드: {kw}\n"
                     f"   설명: {c['과제설명문'][:DESC_CAND]}")
    return "\n".join(lines)


def _parse_scores(text):
    """LLM 출력에서 {"results":[...]} 를 견고하게 파싱 → {id: (score, reason)}."""
    out = {}
    m = re.search(r'\{.*"results".*\}', text, re.DOTALL)
    blob = m.group(0) if m else text
    try:
        data = json.loads(blob)
        items = data.get("results", []) if isinstance(data, dict) else data
    except Exception:
        # 폴백: 개별 객체 단위 정규식 추출
        items = [{"id": int(mm.group(1)), "score": int(mm.group(2)), "reason": mm.group(3)}
                 for mm in re.finditer(
                     r'\{\s*"id"\s*:\s*(\d+)\s*,\s*"score"\s*:\s*(\d+)\s*,\s*"reason"\s*:\s*"([^"]*)"',
                     text)]
    for it in items:
        try:
            i = int(it["id"])
            s = max(0, min(100, int(round(float(it["score"])))))
            out[i] = (s, str(it.get("reason", "")).strip()[:120])
        except (KeyError, ValueError, TypeError):
            continue
    return out


def llm_score(demand, cands):
    """후보(cands: [{id,과제명,과제설명문,키워드}], id=1..K)를 LLM_BATCH 단위로 평가.
    반환 {id: (score, reason)} — 누락분은 score=0."""
    scores = {}
    for b in range(0, len(cands), LLM_BATCH):
        batch = cands[b:b + LLM_BATCH]
        user = f"{_demand_block(demand)}\n\n{_cand_block(batch)}\n\n{_GUIDE}"
        msgs = [{"role": "system", "content": _SYS},
                {"role": "user", "content": user}]
        try:
            out = stream_explanation(msgs, max_tokens=LLM_MAXTOK, temperature=0.0, top_p=1.0)
            parsed = _parse_scores(out)
        except Exception as e:
            print(f"      ! LLM 배치 실패({b}): {e}")
            parsed = {}
        for c in batch:
            scores[c["id"]] = parsed.get(c["id"], (0, "(LLM 응답 누락)"))
    return scores


# ---- 범용 적합도 평가(방향 무관: 과제-기업, 기업-과제, 기업-특허, 수요기술-과제 등) ----
_SYS_DOC = (
    "당신은 기술사업화 매칭을 평가하는 전문가입니다. "
    "기준이 되는 문서(기업·국가 R&D 과제·기술 수요·특허 등)와 후보 문서들 사이의 "
    "실질적 적합도를 냉정하게 평가합니다.")


def _guide_doc(src_label, cand_label):
    return (
        f"각 [{cand_label}] 후보가 위 [{src_label}]와(과) 기술적으로 얼마나 적합·관련되는지 "
        "0~100 정수로 평가하세요.\n"
        "  90-100: 핵심 기술/제품이 직접 부합하는 매우 높은 적합\n"
        "  70-89 : 핵심 기술·응용 분야가 상당 부분 부합\n"
        "  40-69 : 일부 요소만 관련(부분 적합)\n"
        "  10-39 : 분야만 유사하거나 약하게 관련\n"
        "  0-9   : 사실상 무관\n"
        "키워드 표면 일치가 아니라 기술 내용·적용 관점의 실질 적합도를 보세요.\n"
        "반드시 아래 JSON 형식으로만, 모든 후보에 대해 답하세요(reason 은 40자 이내 한국어):\n"
        '{"results": [{"id": <후보번호>, "score": <0-100 정수>, "reason": "<근거>"}, ...]}'
    )


def _src_block(text, label):
    return f"[{label}]\n{text}"


def _cand_block_doc(cands, label, maxlen=DESC_CAND):
    lines = [f"[평가 대상 {label} 목록]"]
    for c in cands:
        lines.append(f"{c['id']}. {str(c.get('text', '')).strip()[:maxlen]}")
    return "\n".join(lines)


def llm_score_doc(source, cands, src_label="기준 문서", cand_label="후보", cand_maxlen=DESC_CAND):
    """방향 무관 범용 적합도 평가(0~100).

    source     : 기준측 문서 텍스트(문자열)
    cands      : [{id, text}] — id=1..K, text=후보 문서 텍스트
    src_label  : 기준측 엔티티 유형명(프롬프트 표시용)
    cand_label : 대상측 엔티티 유형명(프롬프트 표시용)
    cand_maxlen: 후보 텍스트 절단 길이

    예) llm_score_doc(과제문서, 기업후보들, '국가 R&D 과제', '기업')
        llm_score_doc(기업문서, 과제후보들, '기업', '국가 R&D 과제')
        llm_score_doc(기업문서, 특허후보들, '기업', '특허')
        llm_score_doc(수요문서, 과제후보들, '수요기술', '국가 R&D 과제')

    반환 {id: (score, reason)} — 누락분은 score=0."""
    scores = {}
    src = str(source or "").strip()
    guide = _guide_doc(src_label, cand_label)
    for b in range(0, len(cands), LLM_BATCH):
        batch = cands[b:b + LLM_BATCH]
        user = (f"{_src_block(src, src_label)}\n\n"
                f"{_cand_block_doc(batch, cand_label, cand_maxlen)}\n\n{guide}")
        msgs = [{"role": "system", "content": _SYS_DOC},
                {"role": "user", "content": user}]
        try:
            out = stream_explanation(msgs, max_tokens=LLM_MAXTOK, temperature=0.0, top_p=1.0)
            parsed = _parse_scores(out)
        except Exception as e:
            print(f"      ! LLM 배치 실패({b}): {e}")
            parsed = {}
        for c in batch:
            scores[c["id"]] = parsed.get(c["id"], (0, "(LLM 응답 누락)"))
    return scores


# ---- 상세 추천 근거(4섹션) — build_messages/stream_explanation 재사용 + 수요기술 사양 근거 추가 ----
EX_FMT = ["연관성", "수요기술 사양 적합성", "추천 과제의 우수성", "유사 사례 및 실적"]
_EX_GUIDE = {
    "연관성": ("company.description(해당 기업의 사업영역·보유역량·특성)과 company.수요기술명·"
            "수요기술_내용(이 기업이 '필요로 하는'(아직 보유하지 않은) 기술의 특성)을 함께 고려하여, "
            "이 기업의 특성에 비추어 왜 이 수요기술이 필요한지 맥락을 짚고, 그 수요기술과 과제의 "
            "목표·내용 사이의 기술적 연관성을 설명. "
            "수요기술은 기업이 보유한 기술이 아니라 확보하려는 기술이므로 '기업이 이미 ~기술을 "
            "보유/개발했다'거나 '~공정·~소재를 보유하고 있다'고 단정하지 말 것(수요기술 내용을 "
            "기업의 현재 사업·역량으로 서술 금지). company.description 이 비어 있거나 desc_ok=0 이면 "
            "수요기술 정보를 우선 사용하되, 기업이 보유하지 않은 역량을 임의로 가정하지 말 것"),
    "수요기술 사양 적합성": ("company.수요기술_사양의 정량·정성 요구(수치·기준·성능)를 과제가 충족 또는 "
                       "근접하는지, 어떤 항목이 부합하고 어떤 항목이 미흡/불확실한지 구체적으로 평가"),
    "추천 과제의 우수성": "과제의 유망성·연구성과(논문/특허)·수행기관 역량 등 추천 과제의 강점",
    "유사 사례 및 실적": "과제의 논문·특허 등 실적이 수요 해결에 주는 시사점(없으면 생략)",
}


def build_demand_payload(demand, proj):
    """수요기업↔과제 맥락의 build_messages 호환 payload 구성(수요기술 사양 포함).

    수요기술은 기업이 '확보하려는'(미보유) 기술이므로, 기업설명문이 있을 때만
    company.description 으로 쓰고 없으면 비워 둔다(수요기술을 보유역량으로 오인 방지).
    """
    comp_desc = str(demand.get("기업설명문") or "").strip()
    has_cdesc = bool(comp_desc)
    company = {
        "company_id": "", "name": demand.get("기업명", ""),
        "description": comp_desc,
        "desc_issue": str(demand.get("desc_issue", "") or "") if has_cdesc else "",
        "desc_ok": int(demand.get("desc_ok", 1)) if has_cdesc else 1,
        "수요기술명": demand.get("수요기술명", ""),
        "수요기술_내용": demand.get("수요기술 내용", ""),
        "수요기술_사양": demand.get("수요기술 사양", ""),
        "예상_적용_제품_및_서비스": demand.get("예상 적용 제품 및 서비스", ""),
        "keyword": list(demand.get("keywords", []))[:30],
        "purpose": [x for x in (demand.get("수요기술명", ""),
                                demand.get("예상 적용 제품 및 서비스", "")) if x],
        "patent_matching_related": [], "has_patent_matching_related": False,
        "patent_count": None, "company_patent": [],
        "conduct_list_project": [], "has_conduct_project": False, "region": "",
    }
    paper = (proj.get("논문명") or [])[:3]
    patent = (proj.get("특허명") or [])[:3]
    project = {
        "project_id": proj.get("pid", ""), "title": proj.get("과제명", ""),
        "description": (proj.get("설명", "") or "")[:1200],
        "project_score": proj.get("유망성", ""), "cosine_distance": "",
        "keyword": (proj.get("키워드") or [])[:15],
        "과제수행기관": proj.get("수행기관", ""),
        "총연구비_상위비율": proj.get("총연구비_상위비율"),
        "논문건수_상위비율": proj.get("논문건수_상위비율"),
        "특허건수_상위비율": proj.get("특허건수_상위비율"),
    }
    return {
        "company": company, "project": project,
        "conduct_list_company": [], "has_conduct_company": False, "region_relation": "",
        "related_research": {"paper": paper, "patent": patent},
        "has_paper": len(paper) > 0, "has_patent": len(patent) > 0,
        "paper_list_count": int(proj.get("논문건수") or 0),
        "patent_list_count": int(proj.get("특허건수") or 0),
        "output_requirements": {
            "language": "ko", "format": EX_FMT,
            "section_sentence_range": "각 섹션 3~4문장",
            "section_guide": _EX_GUIDE,
            "must_cover_수요기술_사양": True,
            "forbidden": ["예시", "참고", "제출", "메타"],
        },
    }


def gen_explanation(demand, proj):
    """4섹션 상세 추천 근거를 생성해 한 셀 텍스트로 반환(temperature=0.2)."""
    msgs = build_messages(build_demand_payload(demand, proj), direction="company")
    out = stream_explanation(msgs, max_tokens=1400, temperature=0.2, top_p=0.9,
                             expected_keys=EX_FMT)
    secs = parse_sections(out, tuple(EX_FMT))
    parts = [f"[{k}] {secs.get(k, '').strip()}" for k in EX_FMT if secs.get(k, "").strip()]
    return "\n\n".join(parts) if parts else out.strip()


# =====================================================================
# [4] 기업별 탭(시트) 분리 저장 — 셀 wrap 으로 상세근거를 여러 줄로 표시
# =====================================================================
WIDTHS = {                                  # 컬럼별 너비(글자 수). 미지정은 DEFAULT_W.
    "번호": 6, "기업명": 16, "기업임베딩매칭": 12, "기술번호": 12,
    "수요기술명": 34, "키워드": 30, "rank": 5, "LLM점수": 7,
    "LLM판단근거": 45, "과제고유번호": 14, "과제명": 40, "과제수행기관": 22,
    "유사도_과제코사인": 12, "유사도_기업코사인": 12, "유망성점수": 9,
    "추천근거_상세": 70, "과제설명문": 60, "기업설명문_사용": 10, "추천근거_상세_보완": 70,
}
DEFAULT_W = 16
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
HEAD_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
HEAD_FONT = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")


def safe_sheet_name(name, used):
    """엑셀 시트명 제약(31자, []:*?/\\ 금지, 공백/중복) 처리."""
    s = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name)).strip() or "무명"
    s = s[:31]
    base, i = s, 1
    while s.lower() in used:                 # 중복이면 접미사 _1, _2 …
        suf = f"_{i}"
        s = base[:31 - len(suf)] + suf
        i += 1
    used.add(s.lower())
    return s


def split_file(src, dst):
    """src xlsx 를 기업별 시트(같은 기업의 여러 행은 한 시트)로 분리 저장. 모든 셀 wrap."""
    df = pd.read_excel(src)
    wb = Workbook()
    wb.remove(wb.active)
    cols = list(df.columns)
    used = set()
    companies = list(dict.fromkeys(df["기업명"].astype(str).tolist()))
    for comp in companies:
        sub = df[df["기업명"].astype(str) == comp]
        ws = wb.create_sheet(safe_sheet_name(comp, used))
        for c, col in enumerate(cols, 1):                    # 헤더
            hc = ws.cell(row=1, column=c, value=col)
            hc.font, hc.alignment, hc.fill = HEAD_FONT, HEAD_ALIGN, HEAD_FILL
        for r, (_, row) in enumerate(sub.iterrows(), 2):     # 데이터
            for c, col in enumerate(cols, 1):
                v = row[col]
                dc = ws.cell(row=r, column=c, value="" if pd.isna(v) else v)
                dc.alignment = WRAP
        for c, col in enumerate(cols, 1):                    # 너비
            ws.column_dimensions[get_column_letter(c)].width = WIDTHS.get(col, DEFAULT_W)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    wb.save(dst)
    print(f"  → '{dst}' 저장 (기업 탭 {len(companies)}개, 총 {len(df)}행)")


# =====================================================================
# [5] 표준 산출물: 15컬럼 정리본 + 기업별 탭본
# =====================================================================
ORDER = ["번호", "기업명", "기술번호", "수요기술명", "키워드", "과제고유번호", "과제명",
         "과제수행기관", "LLM점수", "LLM판단근거", "추천근거_상세", "과제설명문",
         "유사도_과제코사인", "유사도_기업코사인", "유망성점수"]


def clean_flat(src, dst):
    """src xlsx → ORDER 순서의 정리본(dst). '추천근거_상세_보완'이 있으면 그것으로 승격."""
    df = pd.read_excel(src)
    if "추천근거_상세_보완" in df.columns:
        if "추천근거_상세" in df.columns:
            df = df.drop(columns=["추천근거_상세"])
        df = df.rename(columns={"추천근거_상세_보완": "추천근거_상세"})
    df = df[[c for c in ORDER if c in df.columns]]
    df.to_excel(dst, index=False)
    return df


def make_deliverables(assignee, src=None):
    """담당자 최종추천 → 정리본(_보완) + 기업별 탭본(_보완_기업별) 생성."""
    flat = os.path.join(OUT_DIR, f"COMPA_{assignee}_최종추천_보완.xlsx")
    tabs = os.path.join(OUT_DIR, f"COMPA_{assignee}_최종추천_보완_기업별.xlsx")
    if src is None:                         # 명시 안 하면 기존 정리본/파이프라인 산출물 사용
        src = flat if os.path.exists(flat) else final_xlsx_path(assignee)
    if not os.path.exists(src):
        raise FileNotFoundError(f"원본 없음: {src}")
    clean_flat(src, flat)
    split_file(flat, tabs)
    return flat, tabs


# =====================================================================
# [6] 진입점: 시트 로드 → (담당자별) 키워드 → SBERT → LLM 적합도 → 상세근거 → 산출
# =====================================================================
XLSX = os.path.join(DATA_DIR, "COMPA_진성수요_원본.xlsx")   # 입력(구글시트 다운로드본, 수정 안 함)
ASSIGNEE_COL_IDX = 1                  # 담당자 컬럼(2번째, 헤더 'Unnamed: 1')
DEFAULT_ASSIGNEE = "이중연"           # 기본 대상 담당자
HEADER_ROW = 2                        # pd.read_excel header 행(0-base): 상단 제목/병합 2줄 skip

PROJECT_EMB = os.path.join(DATA_DIR, "public_RnD_embeddings_pro_260601_with_desc.pkl")  # 과제 임베딩
COMPANY_EMB = os.path.join(DATA_DIR, "company_embeddings_pro_260514_with_desc.pkl")     # 기업 설명문(상세근거용)
PROJECT_META = os.path.join(DATA_DIR, "project_match_data_260612.pkl")                  # 과제수행기관/논문/특허 성과
MODEL_DIR = os.path.join(DATA_DIR, "pro-sroberta")

TOPK = 200                 # 재랭킹 후 LLM 평가 후보 수 (= 합집합→재랭커 통과 수)
FINAL = 5                  # 최종 추천 수 (= 상세근거 생성 대상)
KW_MIN, KW_MAX = 10, 20    # 추출 키워드 개수 범위

# --- 합집합(임베딩 ∪ BM25) + 크로스인코더 재랭킹 파라미터 ---
N_EMB = int(os.environ.get("COMPA_N_EMB", "1500"))       # 임베딩(SBERT) 후보 수
N_BM = int(os.environ.get("COMPA_N_BM", "1500"))         # BM25 후보 수
RERANKER_MODEL = os.environ.get("COMPA_RERANKER", "BAAI/bge-reranker-v2-m3")
BM25_K1, BM25_B = 1.5, 0.75
_BM25_TOK = re.compile(r'[가-힣a-zA-Z0-9]{2,}')


# 산출물/체크포인트는 담당자(a)별로 OUT_DIR 아래 분리(번호가 담당자별 1부터여도 충돌 없음).
def kw_xlsx_path(a): return os.path.join(OUT_DIR, f"COMPA_{a}_키워드.xlsx")        # 키워드 추출 결과
def final_xlsx_path(a): return os.path.join(OUT_DIR, f"COMPA_{a}_최종추천.xlsx")   # 최종 추천(xlsx)
def final_pkl_path(a): return os.path.join(OUT_DIR, f"COMPA_{a}_최종추천.pkl")     # 최종 추천(pkl)
def kw_ckpt_path(a): return os.path.join(OUT_DIR, f"compa2stage_{a}_keywords_ckpt.json")   # {번호: [키워드]}
def ex_ckpt_path(a): return os.path.join(OUT_DIR, f"compa2stage_{a}_explain_ckpt.json")    # {번호::과제: text}
def llm_ckpt_path(a):                                                # {번호: {id: [score,reason]}}
    # 합집합→재랭커 후보 기준(옛 SBERT/2단계 체크포인트와 후보 위치가 달라 별도 파일)
    return os.path.join(OUT_DIR, f"compa2stage_{a}_llm_scores_union_ckpt.json")


# ---- LLM 키워드 추출 ----
_KW_SYS = ("당신은 기술 문서에서 핵심 기술 키워드를 추출하는 전문가입니다. "
           "기업의 기술 수요 설명에서 검색·매칭에 유용한 구체적 기술 용어만 선별합니다.")
_KW_GUIDE = (
    "위 [기업 기술 수요]의 수요기술명·수요기술 내용·수요기술 사양에서 핵심 기술 키워드를 "
    f"{KW_MIN}~{KW_MAX}개 추출하세요.\n"
    "  - 소재/공정/성분/기능/대상 등 구체적 기술 명사·전문용어 위주(한글·영문 혼용 허용)\n"
    "  - 일반어(기술, 제품, 개발, 공정, 기반, 적용 등)·수식어·단위·수치는 제외\n"
    "  - 핵심 영문 전문용어(예: Bioavailability, Platycodin D)는 그대로 보존\n"
    "  - 가능한 한 단일/복합 명사 형태의 짧은 키워드로\n"
    "반드시 아래 JSON 형식으로만 답하세요:\n"
    '{"keywords": ["키워드1", "키워드2", ...]}'
)


def _kw_block(d):
    return ("[기업 기술 수요]\n"
            f"- 수요기술명: {d.get('수요기술명','')}\n"
            f"- 수요기술 내용: {d.get('수요기술 내용','')}\n"
            f"- 수요기술 사양: {d.get('수요기술 사양','')}")


def _parse_keywords(text):
    """LLM 출력에서 {"keywords":[...]} 를 견고하게 파싱 → 키워드 리스트(중복/공백 제거)."""
    m = re.search(r'\{.*"keywords".*\}', text, re.DOTALL)
    blob = m.group(0) if m else text
    try:
        data = json.loads(blob)
        items = data.get("keywords", []) if isinstance(data, dict) else data
    except Exception:                       # 폴백: 배열 내용을 정규식으로 추출
        mm = re.search(r'"keywords"\s*:\s*\[(.*?)\]', text, re.DOTALL)
        items = re.findall(r'"([^"]+)"', mm.group(1)) if mm else []
    seen, out = set(), []
    for it in items:
        v = str(it).strip().strip('"').strip()
        if v and v.casefold() not in seen:
            seen.add(v.casefold())
            out.append(v)
    return out[:KW_MAX]


def extract_keywords(demand):
    """수요기술명/내용/사양 → LLM 으로 핵심 기술 키워드 추출(부족 시 규칙기반으로 보충)."""
    msgs = [{"role": "system", "content": _KW_SYS},
            {"role": "user", "content": f"{_kw_block(demand)}\n\n{_KW_GUIDE}"}]
    try:
        kws = _parse_keywords(stream_explanation(msgs, max_tokens=600, temperature=0.0, top_p=1.0))
    except Exception as e:
        print(f"      ! 키워드 추출 LLM 실패: {e}")
        kws = []
    if len(kws) < KW_MIN:                    # 폴백: 규칙기반 명사추출로 보충
        seen = {k.casefold() for k in kws}
        for col in ("수요기술명", "수요기술 내용", "수요기술 사양"):
            for v in normalize_user_input(demand.get(col, "")):
                if v.casefold() not in seen:
                    seen.add(v.casefold())
                    kws.append(v)
                if len(kws) >= KW_MAX:
                    break
    return kws[:KW_MAX]


# ---- 범용 문서 키워드 추출(기업/과제/기술/특허/기술수요 등) ----
_KW_DOC_SYS = (
    "당신은 기술사업화 문서에서 핵심 기술 키워드를 추출하는 전문가입니다. "
    "기업 소개·R&D 과제·기술 설명·특허 명세·기술 수요 등 어떤 유형의 문서가 들어오더라도, "
    "문서의 기술적 실체를 대표하고 검색·매칭에 유용한 구체적 기술 용어만 선별합니다.")
_KW_DOC_GUIDE = (
    "위 [문서]에서 이 문서의 기술적 실체를 대표하는 핵심 기술 키워드를 "
    f"{KW_MIN}~{KW_MAX}개 추출하되, 문서를 가장 잘 대표하는 중요한 키워드부터 순서대로 나열하세요.\n"
    "  - 소재/공정/성분/구조/기능/성능/대상/응용분야 등 구체적 기술 명사·전문용어 위주(한글·영문 혼용 허용)\n"
    "  - 문서 유형에 상관없이, 검색·매칭에 실제로 도움이 되는 식별력 있는 용어만 선별\n"
    "  - 핵심 영문 전문용어·약어·물질명(예: Bioavailability, Platycodin D, CRISPR, OLED)은 그대로 보존\n"
    "  - 특허의 청구항 상용문구(예: \"~하는 것을 특징으로 하는\", \"상기\")나 행정 표현은 키워드로 삼지 말 것\n"
    "  - 가능한 한 단일/복합 명사 형태의 짧은 키워드로, 동일 개념은 하나로 통합\n"
    "아래 불용어(및 이를 포함한 일반적 표현), 수식어·단위·수치·날짜·기관명/사업명은 키워드에서 제외하세요:\n"
    "  기술, 제품, 개발, 연구, 시스템, 방법, 장치, 기반, 적용, 활용,\n"
    "  산업, 방안, 구축, 확보, 강화, 개선, 향상, 최적화, 고도화, 실증,\n"
    "  사업화, 상용화, 지원, 관리, 운영, 도입, 구현, 설계, 제작, 제조,\n"
    "  생산, 분석, 평가, 검증, 관련, 분야, 요소, 과정, 프로세스, 통합,\n"
    "  자동화, 혁신, 융합, 차세대, 첨단, 기타, 제조업, 사업, 회사, 업체,\n"
    "  서비스, 판매, 업무, 제공, 이용, 사용, 처리, 수행, 부문, 종류,\n"
    "  형태, 방식, 절차, 대상, 범위, 내용, 항목, 종목, 품목, 물품,\n"
    "  물자, 시설, 설비, 장비, 기기, 기계, 도구, 용품, 자재, 재료,\n"
    "  원료, 부품, 소재, 물질, 성분, 구조, 형식, 유형, 종별, 구분,\n"
    "  분류, 목록, 도매업, 부대, 일체, 일반, 임대업, 판매업, 상기, 각호,\n"
    "  공급, 호에, 서비스업, 소매업, 부동산, 매매, 전문, 응용, 목적, 대행업,\n"
    "  신품, 공업, 소매, 가공, 상거래, 도매, 형성, 경영, 유지, 기자재,\n"
    "  단계, 용역, 특수, 작업, 유사, 조립, 대행, 가공업, 무역업, 위,\n"
    "  각항, 부대되는, 사업일체, 실시, 진행, 추진, 시행, 완료, 종료, 시작,\n"
    "  착수, 설립, 설치, 반영, 포함, 제외, 및, 외, 내, 중,\n"
    "  상, 하, 전, 후, 등, 것, 수\n"
    "반드시 아래 JSON 형식으로만, 중요도가 높은 키워드부터 순서대로 답하세요:\n"
    '{"keywords": ["가장 중요한 키워드", "다음 키워드", ...]}'
)


# 가이드가 '키워드로 삼지 말라'고 명시한 청구항/행정 상용문구.
# LLM 가이드는 문장으로 배제를 지시하지만 규칙기반 폴백은 그 지시를 따르지 못하므로,
# 최종 산출에서 두 경로(LLM·폴백) 모두에 일괄 적용해 특허 청구항 상용어 누수를 막는다.
_KW_DOC_BOILERPLATE = {"특징", "상기", "것을", "것", "포함하는", "그것"}


def _doc_block(text):
    return f"[문서]\n{text}"


def extract_keywords_doc(text):
    """기업/과제/기술/특허/기술수요 등 임의의 기술사업화 문서 텍스트(문자열) →
    LLM 으로 핵심 기술 키워드 추출(부족 시 규칙기반으로 보충)."""
    text = str(text or "").strip()
    msgs = [{"role": "system", "content": _KW_DOC_SYS},
            {"role": "user", "content": f"{_doc_block(text)}\n\n{_KW_DOC_GUIDE}"}]
    try:
        kws = _parse_keywords(stream_explanation(msgs, max_tokens=600, temperature=0.0, top_p=1.0))
    except Exception as e:
        print(f"      ! 문서 키워드 추출 LLM 실패: {e}")
        kws = []
    if len(kws) < KW_MIN:                    # 폴백: 규칙기반 명사추출로 보충
        seen = {k.casefold() for k in kws}
        for v in normalize_user_input(text):
            if v.casefold() not in seen:
                seen.add(v.casefold())
                kws.append(v)
            if len(kws) >= KW_MAX:
                break
    # 가이드의 청구항/행정 상용문구 배제를 LLM·폴백 최종 산출에 일괄 적용
    kws = [k for k in kws if k not in _KW_DOC_BOILERPLATE]
    return kws[:KW_MAX]


def build_company_desc_index(cdf):
    """회사 임베딩 DF → {정규화기업명: (기업설명문, desc_ok, desc_issue)}.
    동일 정규화명이 여럿이면 설명문이 더 긴(정보 많은) 행을 채택."""
    if "기업설명문" not in cdf.columns:
        return {}
    idx, best = {}, {}
    names = cdf["한글업체명"].astype(str).values
    descs = cdf["기업설명문"].astype(str).values
    oks = cdf["desc_ok"].values if "desc_ok" in cdf.columns else [1] * len(cdf)
    isss = (cdf["desc_issue"].astype(str).values if "desc_issue" in cdf.columns
            else [""] * len(cdf))
    for i in range(len(cdf)):
        key = norm_name(names[i])
        if not key:
            continue
        d = descs[i].strip()
        if d.lower() == "nan":
            d = ""
        if key in idx and len(d) <= best.get(key, -1):
            continue
        try:
            ok = int(oks[i])
        except Exception:
            ok = 1
        iss = str(isss[i]).strip()
        if iss.lower() == "nan":
            iss = ""
        idx[key] = (d, ok, iss)
        best[key] = len(d)
    return idx


def make_encoder(model):
    """텍스트 → pro-sroberta 임베딩(L2 정규화) 인코더 클로저."""
    def encode(text):
        e = model.encode([text], convert_to_numpy=True, normalize_embeddings=False,
                         show_progress_bar=False)[0].astype(np.float32)
        n = np.linalg.norm(e)
        return e / n if n else e
    return encode


# ---- BM25 (어휘 검색) — 합집합 후보의 한 축 ----
def _bm25_tokens(text):
    """BM25용 토크나이저: 한글/영숫자 2글자 이상."""
    return _BM25_TOK.findall(str(text).lower())


def build_bm25(pname, pdesc, pkw):
    """과제명+설명문+키워드로 BM25 역색인 구축 → {cv, X(csc), idf, Kdoc, N}."""
    from sklearn.feature_extraction.text import CountVectorizer
    n = len(pname)
    docs = [f"{pname[i]} {pdesc[i]} "
            f"{' '.join(pkw[i]) if isinstance(pkw[i], (list, tuple)) else ''}" for i in range(n)]
    cv = CountVectorizer(tokenizer=_bm25_tokens, token_pattern=None,
                         min_df=2, max_features=400000)
    X = cv.fit_transform(docs).tocsc()
    dl = np.asarray(X.sum(axis=1)).ravel().astype(np.float32)
    avgdl = float(dl.mean()) or 1.0
    df = np.diff(X.indptr)
    idf = np.log(1 + (n - df + 0.5) / (df + 0.5)).astype(np.float32)
    kdoc = (BM25_K1 * (1 - BM25_B + BM25_B * dl / avgdl)).astype(np.float32)
    return {"cv": cv, "X": X, "idf": idf, "Kdoc": kdoc, "N": n}


def bm25_scores(bm, query_terms):
    """질의 토큰 → 전체 과제 BM25 점수 배열(길이 N)."""
    cv, X, idf, kdoc, n = bm["cv"], bm["X"], bm["idf"], bm["Kdoc"], bm["N"]
    ids = {cv.vocabulary_.get(w) for w in query_terms}
    ids.discard(None)
    sc = np.zeros(n, dtype=np.float32)
    for t in ids:
        col = X.getcol(t)
        r, tf = col.indices, col.data.astype(np.float32)
        sc[r] += idf[t] * (tf * (BM25_K1 + 1)) / (tf + kdoc[r])
    return sc


# ---- 크로스인코더 재랭커 (lazy 로드, 프로세스당 1회) ----
_reranker = None


def load_reranker():
    global _reranker
    if _reranker is None:
        import torch
        from sentence_transformers import CrossEncoder
        dev = ("cuda" if torch.cuda.is_available()
               else ("mps" if getattr(torch.backends, "mps", None)
                     and torch.backends.mps.is_available() else "cpu"))
        _reranker = CrossEncoder(RERANKER_MODEL, max_length=512, device=dev)
    return _reranker


def retrieve_candidates(demand, kws, corpus, encode, topk):
    """합집합 검색 → 재랭킹 → 상위 topk 후보(전역 인덱스) 반환. (cos1 도 함께 반환)

    ① 임베딩: 키워드 문자열 → pro-sroberta → 과제 임베딩 코사인 상위 N_EMB
    ② BM25 : 수요원문(명사정제)+키워드 → 어휘 상위 N_BM
    ③ 합집합을 크로스인코더로 관련도순 재정렬 → 상위 topk
    """
    M, pname, pdesc = corpus["M"], corpus["pname"], corpus["pdesc"]
    # ① 임베딩 후보 + 전체 코사인(출력/동점처리용)
    cos1 = np.clip(M @ encode(SEP.join(kws)), -1, 1)
    emb_top = np.argsort(-cos1)[:N_EMB]
    # ② BM25 후보
    dtext = " ".join(demand.get(c, "") for c in
                     ("수요기술명", "수요기술 내용", "수요기술 사양", "예상 적용 제품 및 서비스"))
    bmq = _bm25_tokens(" ".join(normalize_user_input(dtext)) + " " + SEP.join(kws))
    bm_top = np.argsort(-bm25_scores(corpus["bm25"], bmq))[:N_BM]
    pool = np.unique(np.concatenate([emb_top, bm_top]))
    # ③ 재랭킹 (수요 원문을 질의로)
    query = (demand.get("수요기술명", "") + " " + demand.get("수요기술 내용", ""))[:300]
    pairs = [[query, (pname[i] + ". " + pdesc[i])[:800]] for i in pool]
    rr = load_reranker().predict(pairs, batch_size=64, show_progress_bar=False)
    cand_idx = pool[np.argsort(-rr)[:topk]]
    return cand_idx, cos1


def load_corpus():
    """과제 임베딩 + 기업 설명문 + 과제 메타 + BM25 색인을 1회 로드(여러 담당자에 재사용)."""
    print("  · 과제 임베딩 로드…", flush=True)
    pdf = pd.read_pickle(PROJECT_EMB)
    corpus = {
        "pid": pdf["과제고유번호"].astype(str).values,
        "pname": pdf["과제명"].astype(str).values,
        "pdesc": pdf["과제설명문"].astype(str).values,
        "pkw": pdf["키워드_리스트"].values,
        "promise": pdf["유망성점수"].values.astype(np.float32),
    }
    M = np.vstack([np.asarray(e, dtype=np.float32) for e in pdf["norm_embed"].values])
    M /= np.linalg.norm(M, axis=1, keepdims=True)          # 행별 L2 정규화
    corpus["M"] = M
    del pdf
    print("  · BM25 색인 구축(어휘 검색)…", flush=True)
    corpus["bm25"] = build_bm25(corpus["pname"], corpus["pdesc"], corpus["pkw"])
    print("  · 기업 설명문 로드(상세근거용)…", flush=True)
    cdf = pd.read_pickle(COMPANY_EMB)
    corpus["cdesc_idx"] = build_company_desc_index(cdf)
    del cdf
    print("  · 과제 메타 로드…", flush=True)
    with open(PROJECT_META, "rb") as f:
        pmeta = pickle.load(f)
    corpus["pmeta"] = pmeta
    corpus["org"] = np.array([str(pmeta.get(p, {}).get("과제수행기관명", "")) for p in corpus["pid"]])
    return corpus


def extract_keywords_for(assignee, records):
    """담당자 수요들의 키워드 추출(체크포인트 재사용) → kw_ckpt 반환 + 키워드 xlsx 저장."""
    ckpt_path = kw_ckpt_path(assignee)
    kw_ckpt = {}
    if os.path.exists(ckpt_path):
        with open(ckpt_path, encoding="utf-8") as f:
            kw_ckpt = json.load(f)
        print(f"      키워드 체크포인트 로드: {len(kw_ckpt)}건")
    kw_rows = []
    for n, rec in enumerate(records, 1):
        demand_no = cell(rec, "번호")
        company = cell(rec, "기업명")
        ck = str(demand_no)
        if ck not in kw_ckpt:                # 미처리분만 LLM 호출 후 즉시 체크포인트 저장
            demand = {c: cell(rec, c) for c in
                      ("수요기술명", "수요기술 내용", "수요기술 사양", "예상 적용 제품 및 서비스")}
            kw_ckpt[ck] = extract_keywords(demand)
            with open(ckpt_path, "w", encoding="utf-8") as f:
                json.dump(kw_ckpt, f, ensure_ascii=False)
        kws = kw_ckpt[ck]
        row = dict(rec)
        row["키워드"] = SEP.join(kws)
        kw_rows.append(row)
        print(f"      [{demand_no}] {company}: 키워드 {len(kws)}개 [{n}/{len(records)}]")
    out = kw_xlsx_path(assignee)
    pd.DataFrame(kw_rows).to_excel(out, index=False)
    print(f"      → '{out}' 저장({len(kw_rows)}행)")
    return kw_ckpt


def match_for(assignee, records, kw_ckpt, corpus, encode, args):
    """담당자 수요들: SBERT 매칭 + LLM 적합도 + 상세근거(체크포인트 재사용) → 최종 xlsx + 후처리."""
    topk, final = args.topk, args.final
    pid, pname, pdesc = corpus["pid"], corpus["pname"], corpus["pdesc"]
    pkw, promise, M = corpus["pkw"], corpus["promise"], corpus["M"]
    pmeta, org = corpus["pmeta"], corpus["org"]
    cdesc_idx = corpus.get("cdesc_idx", {})

    llm_path, ex_path = llm_ckpt_path(assignee), ex_ckpt_path(assignee)
    llm_ckpt, ex_ckpt = {}, {}
    if os.path.exists(llm_path):
        with open(llm_path, encoding="utf-8") as f:
            llm_ckpt = json.load(f)
        print(f"      LLM점수 체크포인트 로드: {len(llm_ckpt)}건")
    if os.path.exists(ex_path):
        with open(ex_path, encoding="utf-8") as f:
            ex_ckpt = json.load(f)
        print(f"      상세근거 체크포인트 로드: {len(ex_ckpt)}건")

    final_rows = []
    for n, rec in enumerate(records, 1):
        demand_no = cell(rec, "번호")
        company = cell(rec, "기업명")
        kws = kw_ckpt[str(demand_no)]
        kw_string = SEP.join(kws)

        # LLM 적합도 평가용 demand (수요기술 기준). 상세근거용(demand_ex)엔 기업설명문을 덧붙임.
        demand = {c: cell(rec, c) for c in
                  ("수요기술명", "수요기술 내용", "수요기술 사양", "예상 적용 제품 및 서비스")}
        demand["기업명"] = company
        demand["keywords"] = kws
        demand_ex = dict(demand)
        cdesc = cdesc_idx.get(norm_name(company))
        if cdesc and cdesc[0]:
            demand_ex["기업설명문"], demand_ex["desc_ok"], demand_ex["desc_issue"] = cdesc

        # 합집합 검색: 임베딩(top N_EMB) ∪ BM25(top N_BM) → 크로스인코더 재랭킹 → 상위 topk 후보
        cand_idx, cos1 = retrieve_candidates(demand, kws, corpus, encode, topk)

        cands = [{"id": j + 1, "과제명": pname[i], "과제설명문": pdesc[i],
                  "키워드": list(pkw[i]) if isinstance(pkw[i], (list, tuple)) else []}
                 for j, i in enumerate(cand_idx)]
        ck = str(demand_no)
        if ck in llm_ckpt:                   # 체크포인트 재사용(재실행 시 LLM 재호출 안 함)
            id2 = {int(k): tuple(v) for k, v in llm_ckpt[ck].items()}
        else:
            id2 = llm_score(demand, cands)
            llm_ckpt[ck] = {str(k): list(v) for k, v in id2.items()}
            with open(llm_path, "w", encoding="utf-8") as f:
                json.dump(llm_ckpt, f, ensure_ascii=False)

        # LLM 점수 내림차순 → 최종 final 개 (동점 시 과제 코사인). id(1..K)→코퍼스 인덱스.
        scored = [(id2.get(j + 1, (0, ""))[0], float(cos1[i]), j, i, id2.get(j + 1, (0, ""))[1])
                  for j, i in enumerate(cand_idx)]
        scored.sort(key=lambda x: (-x[0], -x[1]))

        # 과제명 중복 제거: 동일 과제명은 점수 높은(먼저 오는) 것만 남기고 final 개까지 선정
        seen_titles, selected = set(), []
        for tup in scored:
            tkey = str(pname[tup[3]]).strip().casefold()
            if tkey in seen_titles:
                continue
            seen_titles.add(tkey)
            selected.append(tup)
            if len(selected) >= final:
                break

        for rank, (s, _c1, j, i, reason) in enumerate(selected, 1):
            # 상세 근거는 최종 final 개에 한해서만 생성(--no-explain 이면 생략, 별도 패스 지원)
            ekey = f"{demand_no}::{pid[i]}"
            if not args.no_explain and ekey not in ex_ckpt:
                meta = pmeta.get(pid[i], {}) if isinstance(pmeta, dict) else {}
                proj = {
                    "pid": pid[i], "과제명": pname[i], "설명": pdesc[i],
                    "유망성": round(float(promise[i]), 1), "수행기관": org[i],
                    "키워드": list(pkw[i]) if isinstance(pkw[i], (list, tuple)) else [],
                    "논문명": meta.get("논문명_리스트") or [],
                    "특허명": meta.get("특허명_리스트") or [],
                    "논문건수": meta.get("논문건수", 0), "특허건수": meta.get("특허건수", 0),
                    "총연구비_상위비율": meta.get("총연구비_상위비율"),
                    "논문건수_상위비율": meta.get("논문건수_상위비율"),
                    "특허건수_상위비율": meta.get("특허건수_상위비율"),
                }
                ex_ckpt[ekey] = gen_explanation(demand_ex, proj)[:DESC_OUT]
                with open(ex_path, "w", encoding="utf-8") as f:
                    json.dump(ex_ckpt, f, ensure_ascii=False)
            final_rows.append({
                "번호": demand_no, "기업명": company,
                "기술번호": cell(rec, "기술번호"), "수요기술명": cell(rec, "수요기술명"),
                "키워드": kw_string, "rank": rank,
                "LLM점수": s, "LLM판단근거": reason,
                "과제고유번호": pid[i], "과제명": pname[i], "과제수행기관": org[i],
                "유사도_과제코사인": round(float(cos1[i]), 6),
                "유망성점수": round(float(promise[i]), 4),
                "추천근거_상세": ex_ckpt.get(ekey, ""),
                "과제설명문": pdesc[i][:DESC_OUT],
            })
        best = scored[0][0] if scored else 0
        print(f"      [{demand_no}] {company}: SBERT top{len(cand_idx)} "
              f"→ LLM best={best} → 최종 {len(selected)} [{n}/{len(records)}]")

    out = final_xlsx_path(assignee)
    df_final = pd.DataFrame(final_rows)
    df_final.to_excel(out, index=False)
    df_final.to_pickle(final_pkl_path(assignee))     # 결과를 pkl 로도 저장(재사용/분석 편의)
    print(f"      → '{out}' (+ .pkl) 저장(최종추천 {len(final_rows)}행)")

    # 표준 산출물 자동 생성: 15컬럼 정리본(_보완) + 기업별 탭본(_보완_기업별)
    try:
        flat, tabs = make_deliverables(assignee, src=out)
        print(f"      → 후처리: '{flat}', '{tabs}' 생성")
    except Exception as e:
        print(f"      ! 후처리(정리/탭 분리) 실패: {e}")


def resolve_assignees(xl, acol, args):
    """처리할 담당자 목록 결정: --all(시트 전체) / --assignee(쉼표 구분) / 기본값, 그리고 --exclude."""
    present = []
    for v in xl[acol].astype(str).str.strip():
        if v and v.lower() != "nan" and v not in present:
            present.append(v)
    if args.all:
        targets = present
    elif args.assignee:
        targets = [a.strip() for a in args.assignee.split(",") if a.strip()]
    else:
        targets = [DEFAULT_ASSIGNEE]
    exclude = {e.strip() for e in (args.exclude or "").split(",") if e.strip()}
    return [a for a in targets if a not in exclude]


def run(args):
    """파싱된 args(Namespace)로 전체 파이프라인 실행. CLI(main)·노트북(run_kw) 공통 진입."""
    print("[1/4] 시트 로드 + 담당자 결정…")
    xl = pd.read_excel(XLSX, header=HEADER_ROW)
    acol = xl.columns[ASSIGNEE_COL_IDX]
    assignees = resolve_assignees(xl, acol, args)
    print(f"      대상 담당자({len(assignees)}): {assignees} (대상 컬럼='{acol}', 시트 전체 {len(xl)})")

    print("[2/4] pro-sroberta 로드…")
    model = SentenceTransformer(MODEL_DIR, device="cpu")
    model.eval()
    encode = make_encoder(model)

    corpus = None
    if not args.keywords_only:
        print("[3/4] 과제/기업 임베딩 코퍼스 로드(1회)…")
        corpus = load_corpus()
    else:
        print("[3/4] (keywords-only: 코퍼스 로드 생략)")

    print("[4/4] 담당자별 처리…")
    for ai, assignee in enumerate(assignees, 1):
        target = xl[xl[acol].astype(str).str.strip() == assignee].reset_index(drop=True)
        records = target.to_dict("records")
        if args.limit:
            records = records[:args.limit]
        print(f"  === [{ai}/{len(assignees)}] 담당자='{assignee}' 수요 {len(records)}건 ===")
        kw_ckpt = extract_keywords_for(assignee, records)
        if args.keywords_only:
            continue
        match_for(assignee, records, kw_ckpt, corpus, encode, args)
    print("완료.")


def run_kw(**kwargs):
    """노트북/프로그램에서 키워드 인자로 호출. 예: run_kw(all=True, no_explain=False)."""
    d = dict(assignee="", all=False, exclude="", limit=0, topk=TOPK, final=FINAL,
             no_explain=False, keywords_only=False)
    d.update(kwargs)
    run(argparse.Namespace(**d))


def main():
    ap = argparse.ArgumentParser(description="COMPA 진성수요 ↔ 국가 R&D 과제 매칭")
    ap.add_argument("--assignee", default="", help="대상 담당자(쉼표로 여러 명). 미지정+--all 아니면 기본='이중연'")
    ap.add_argument("--all", action="store_true", help="시트의 모든 담당자 처리(담당자별 파일 분리)")
    ap.add_argument("--exclude", default="", help="제외할 담당자(쉼표 구분). 예: --all --exclude 이중연")
    ap.add_argument("--limit", type=int, default=0, help="담당자별 처리 수요 수 제한(0=전체, 테스트용)")
    ap.add_argument("--topk", type=int, default=TOPK, help="재랭킹 후 LLM 평가 후보 수(기본 200)")
    ap.add_argument("--final", type=int, default=FINAL, help="최종 추천 수")
    ap.add_argument("--no-explain", action="store_true",
                    help="상세근거 생성 생략(매칭/LLM점수까지만). 상세근거는 별도 패스에서 "
                         "다른 모델(예: MV_MLX_MODEL=...Qwen3.5-35B-A3B-4bit)로 채운다.")
    ap.add_argument("--keywords-only", action="store_true",
                    help="키워드 추출 + 키워드 xlsx 저장까지만 수행(매칭/LLM 스킵)")
    run(ap.parse_args())


if __name__ == "__main__":
    main()
